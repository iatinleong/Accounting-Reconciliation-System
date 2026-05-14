import pandas as pd
import json
import os
import re
import warnings
import itertools
from IPython.display import display, HTML
import win32com.client as win32
warnings.filterwarnings('ignore')
pd.set_option('display.float_format', lambda x: '%.2f' % x)
pd.set_option('display.max_columns', None)

import openai
api_key = os.environ.get("OPENAI_API_KEY", "")
client = openai.OpenAI(api_key=api_key)

# ── 1. 掃描資料夾：支援多份銀行對帳單與多份帳務查詢 ──
# 同一底名（不同副檔名）只取「最佳」版本：xlsx > xlsm > xls
# 避免 convert_to_xlsx 轉出的 .xlsx 與原始檔重複載入。
data_dir = r'C:\Users\user\Desktop\專案\會計對帳\資料'
_EXT_PRIORITY = {'.xlsx': 0, '.xlsm': 1, '.xls': 2}

# 先按底名分組，每組只保留優先度最高的副檔名
_candidates = {}   # {底名: (優先度, 完整路徑)}
for fname in os.listdir(data_dir):
    if fname.startswith('~$'): continue
    base, ext = os.path.splitext(fname)
    ext = ext.lower()
    if ext not in _EXT_PRIORITY: continue
    fpath = os.path.join(data_dir, fname)
    pri = _EXT_PRIORITY[ext]
    if base not in _candidates or pri < _candidates[base][0]:
        _candidates[base] = (pri, fpath)

bank_files = {}   # {月份碼: 路徑}
acc_paths  = []   # 所有帳務查詢檔案路徑

for base, (_, fpath) in sorted(_candidates.items()):
    if '銀行對帳單' in base:
        m = re.search(r'銀行對帳單-(\d{5,8})', base)
        if m:
            code = m.group(1)[:5]   # 前5碼 = 民國YYMM
            bank_files[code] = fpath
    elif '帳務查詢' in base:
        acc_paths.append(fpath)

assert bank_files, '找不到銀行對帳單檔案'
assert acc_paths,  '找不到帳務查詢檔案'


# ── 2. 工具函數 ──
def convert_to_xlsx(src_path):
    if src_path.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            load_workbook(src_path, read_only=True).close()
            return src_path
        except Exception:
            pass

    import win32com.client, pythoncom
    pythoncom.CoInitialize()
    dst_path = src_path.rsplit('.', 1)[0] + '.xlsx'
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(src_path)
        wb.SaveAs(dst_path, FileFormat=51)
        wb.Close(False)
        print(f'🔄 已轉換：{os.path.basename(dst_path)}')
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()
    return dst_path


def month_code_to_date(code):
    """民國年月碼 '11501' → '2026/01'"""
    return f'{int(code[:3]) + 1911}/{code[3:5]}'


# ── 3. 轉換所有帳務查詢並合併全部 Sheet（避免每月重複 IO）──
# 若多份帳務查詢的 sheet 名稱相同，用「檔名::sheet名」做唯一鍵，避免互蓋。
_acc_sheets_global = {}
for ap in acc_paths:
    ap_xlsx = convert_to_xlsx(ap)
    src_label = os.path.splitext(os.path.basename(ap_xlsx))[0]  # 去副檔名的檔名
    sheets = pd.read_excel(ap_xlsx, sheet_name=None)
    for sheet_name, df in sheets.items():
        unique_key = f'{src_label}::{sheet_name}' if sheet_name in _acc_sheets_global else sheet_name
        _acc_sheets_global[unique_key] = df
    print(f'✅ 載入 {os.path.basename(ap_xlsx)}，{len(sheets)} 個 Sheet')

print(f'✅ 會計帳合計 {len(_acc_sheets_global)} 個 Sheet：{list(_acc_sheets_global.keys())}')
print(f'✅ 找到 {len(bank_files)} 份銀行對帳單：{sorted(bank_files.keys())}')


# ── 4. 載入與預處理 ──
def load_and_preprocess(target_date, bank_path_param=None, acc_sheets_param=None):
    _bank_path = bank_path_param
    _all_sheets = acc_sheets_param if acc_sheets_param is not None else {}

    # --- 銀行端 ---
    df_bank = pd.read_excel(_bank_path, skiprows=5)
    df_bank = df_bank.dropna(subset=['交易日期'])
    df_bank['存入金額'] = pd.to_numeric(df_bank['存入金額'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)
    df_bank['支出金額'] = pd.to_numeric(df_bank['支出金額'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0)

    target_month = target_date[:7]
    df_bank_all = df_bank.copy()
    df_bank_all['_date_str'] = pd.to_datetime(df_bank_all['交易日期'], errors='coerce').dt.strftime('%Y/%m')
    df_bank_all = df_bank_all[df_bank_all['_date_str'] == target_month].drop(columns=['_date_str'])
    df_bank_all['Bank_index'] = df_bank_all.index.astype(str)

    # --- 日內沖銷 Pre-processing ---
    df_bank_all['_date'] = pd.to_datetime(df_bank_all['交易日期'], errors='coerce').dt.date
    to_remove = set()
    reversal_log = []
    reversals = df_bank_all[df_bank_all['支出金額'] < 0]
    for r_idx, r_row in reversals.iterrows():
        rev_amt = abs(r_row['支出金額'])
        matches = df_bank_all[
            (df_bank_all['_date'] == r_row['_date']) &
            (df_bank_all['支出金額'] == rev_amt) &
            (~df_bank_all.index.isin(to_remove))
        ]
        if not matches.empty:
            orig_idx = matches.index[0]
            to_remove.add(r_idx)
            to_remove.add(orig_idx)
            reversal_log.append({
                'orig': df_bank_all.loc[orig_idx].copy(),
                'chong': r_row.copy(),
                'amount': rev_amt
            })
            print(f"  🔄 日內沖銷移除：Bank_index={df_bank_all.loc[orig_idx,'Bank_index']} [{df_bank_all.loc[orig_idx,'摘要']}] + [{r_row['摘要']}]  各 {rev_amt:,.0f}")
    df_bank_all = df_bank_all[~df_bank_all.index.isin(to_remove)].drop(columns=['_date'])

    bank_in  = df_bank_all[df_bank_all['存入金額'] > 0].copy()
    bank_out = df_bank_all[df_bank_all['支出金額'] > 0].copy()

    # --- 會計端 ---
    required_acc_columns = [
        '憑證創建人', '憑證編號', '憑證行編號', '憑證類型', '業務參考',
        '會計期間', '業務日期', '科目代碼', '描述.1', '業務貨幣代碼',
        '業務金額', '本位幣金額', '第二本位幣/報表金額', '借方/貸方',
        '商部別 分析代碼', '定存單號/銀行與帳號 分析代碼', '對象別 分析代碼'
    ]

    df_acc_list = []
    for sheet_name, df_sheet in _all_sheets.items():
        actual_cols = [c for c in df_sheet.columns if c in required_acc_columns or c == '描述.1' or c == '業務金額' or c == '業務日期']
        if '業務金額' not in df_sheet.columns or '業務日期' not in df_sheet.columns:
            continue
        df_sheet = df_sheet[actual_cols].copy()
        df_sheet['來源Sheet'] = sheet_name
        df_acc_list.append(df_sheet)
        print(f'  ✅ 讀入 [{sheet_name}]：{len(df_sheet)} 筆')

    df_acc = pd.concat(df_acc_list, ignore_index=True)

    def parse_amount(s):
        s = str(s).strip().replace(',', '')
        if s.startswith('(') and s.endswith(')'):
            s = '-' + s[1:-1]
        return pd.to_numeric(s, errors='coerce')
    df_acc['業務金額'] = df_acc['業務金額'].apply(parse_amount).fillna(0)

    try:
        df_acc['業務日期_格式化'] = pd.to_datetime(df_acc['業務日期'], errors='coerce').dt.strftime('%Y/%m/%d')
        if df_acc['業務日期_格式化'].isna().all():
            raise ValueError()
    except Exception:
        df_acc['業務日期_格式化'] = pd.to_datetime(
            pd.to_numeric(df_acc['業務日期'], errors='coerce'),
            origin='1899-12-30', unit='D'
        ).dt.strftime('%Y/%m/%d')

    df_acc_all = df_acc[df_acc['業務日期_格式化'].str.startswith(target_month, na=False)].copy()
    df_acc_all['Acc_index'] = df_acc_all.index.astype(str)

    acc_in  = df_acc_all[df_acc_all['業務金額'] > 0].copy()
    acc_out = df_acc_all[df_acc_all['業務金額'] < 0].copy()

    return bank_in, bank_out, acc_in, acc_out, reversal_log


# ── 5. 關鍵字與 LLM 輔助函數 ──
def extract_keywords(text):
    text = str(text).strip()
    if not text or text.lower() == 'nan':
        return []
    tokens = re.findall(r'[A-Za-z0-9]{3,}|[一-鿿]{2,}', text)
    return [t.upper() for t in tokens]

def _cjk_substrings(token, min_len=2):
    """從中文 token 產生所有長度 >= min_len 的子字串（由長到短），用於滑窗比對。"""
    for sub_len in range(len(token), min_len - 1, -1):
        for start in range(len(token) - sub_len + 1):
            yield token[start:start + sub_len]

def keyword_match(bank_memo, acc_desc1):
    bank_tokens = extract_keywords(bank_memo)
    acc_text = str(acc_desc1).upper()
    for token in bank_tokens:
        if token in acc_text:
            return True, token
        # 中文長 token 做子字串滑窗（處理「台灣曼茲科技股」vs「台灣曼茲10月份」的情況）
        if len(token) >= 4 and all('一' <= c <= '鿿' for c in token):
            for sub in _cjk_substrings(token, min_len=2):
                if sub in acc_text:
                    return True, sub
    bank_text = str(bank_memo).upper()
    for token in extract_keywords(acc_desc1):
        if len(token) >= 4 and token in bank_text:
            return True, token
        # 反向也做滑窗
        if len(token) >= 4 and all('一' <= c <= '鿿' for c in token):
            for sub in _cjk_substrings(token, min_len=2):
                if sub in bank_text:
                    return True, sub
    return False, ''

def check_memo_match(bank_memo, acc_desc_extra):
    b_text = str(bank_memo).strip()
    a_text = str(acc_desc_extra).strip()
    if not b_text or b_text.lower() in ['nan']: return True
    if a_text and a_text.lower() not in ['nan']:
        if (b_text in a_text) or (a_text in b_text) or len(set(b_text) & set(a_text)) >= 2: return True
    return False

def llm_pick_best_candidate(bank_row, candidates_df, mode):
    if '描述.1' in candidates_df.columns:
        if candidates_df['描述.1'].astype(str).str.contains('銀行存款－台北富邦銀行', na=False).all():
            print(f"      ⏭️ 所有候選描述均為銀行存款，LLM 無法分辨，留待批次核銷")
            return None

    prompt = f"""你是專業會計。有一筆銀行{('收款' if mode=='Income' else '付款')}，與其金額相同的會計帳有 {len(candidates_df)} 筆候選紀錄。
請根據銀行的「附言」與會計帳的「備註描述(描述.1)」，挑出真正對應的「唯一 1 筆」會計帳。
【目標銀行紀錄】
 Bank_index: {bank_row['Bank_index']}, 附言: {bank_row['附言']}
【多筆會計候選人】
{candidates_df[['Acc_index', '描述.1']].to_json(orient='records', force_ascii=False)}
請只回傳最符合的該筆 Acc_index 字串。若都不吻合則回傳 None。"""
    try:
        resp = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0
        )
        ans = resp.choices[0].message.content.strip()
        if ans in candidates_df['Acc_index'].values: return ans
    except Exception:
        pass
    return None


def llm_pick_best_bank_candidate(acc_row, bank_candidates_df, mode):
    """給定 1 筆會計，從多筆同金額銀行中挑出最吻合的。"""
    prompt = f"""你是專業會計。有一筆會計{('收款' if mode=='Income' else '付款')}，與其金額相同的銀行記錄有 {len(bank_candidates_df)} 筆候選。
請根據會計帳的「備註描述(描述.1)」與銀行的「附言」，挑出真正對應的「唯一 1 筆」銀行記錄。
【目標會計紀錄】
 Acc_index: {acc_row['Acc_index']}, 描述.1: {acc_row.get('描述.1','')}, 憑證編號: {acc_row.get('憑證編號','')}
【多筆銀行候選人】
{bank_candidates_df[['Bank_index', '附言', '摘要']].to_json(orient='records', force_ascii=False)}
請只回傳最符合的該筆 Bank_index 字串。若都不吻合則回傳 None。"""
    try:
        resp = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
            temperature=0
        )
        ans = resp.choices[0].message.content.strip()
        if ans in bank_candidates_df['Bank_index'].values: return ans
    except Exception:
        pass
    return None


COIN_MIN_AMT   = 1_000_000
COIN_TOLERANCE = 2_000

def coin_fill(pool, amt_col, target):
    selected, rem = [], int(round(target))
    for bamt_f in sorted(pool[amt_col].unique(), reverse=True):
        bamt = int(round(bamt_f))
        if bamt <= 0 or bamt > rem: continue
        cands = pool[
            (pool[amt_col].apply(lambda x: int(round(x))) == bamt) &
            (~pool['Bank_index'].isin(selected))
        ]
        n = min(rem // bamt, len(cands))
        if n > 0:
            selected.extend(cands.iloc[:n]['Bank_index'].tolist())
            rem -= n * bamt
        if rem == 0:
            break
    return selected, rem

def evaluate_combination_with_llm(bank_rows, acc_rows, match_type, mode):
    if '描述.1' in acc_rows.columns:
        if acc_rows['描述.1'].astype(str).str.contains('銀行存款－台北富邦銀行', na=False).all():
            return True, "會計描述為銀行存款類，無需LLM，數學完全吻合自動通過"
    return True, f"{mode}端數學完全吻合，自動通過"


# ── 6. 對帳引擎 ──
def reconcile_engine(df_bank, df_acc, mode='Income', pending=None):
    if pending is None: pending = []
    label = "收入" if mode == "Income" else "支出"
    print(f"\n{'='*16} 🚀 [{label}] 開始對帳 {'='*16}")
    rem_b, rem_a = df_bank.copy(), df_acc.copy()
    history = []
    bank_amt_col = '存入金額' if mode == 'Income' else '支出金額'

    # Step 1: 1對1（兩輪：先關鍵字優先，再金額唯一）
    print(f"\n⏳ [Step 1-A] 關鍵字優先輪...")
    # 輪一：只要銀行附言與會計描述有關鍵字命中，不管候選數多少，立即配對
    # 避免無關鍵字的銀行筆因「金額唯一」搶走有語意對應的會計筆
    for b_idx, b_row in df_bank.iterrows():
        if b_idx not in rem_b.index: continue
        b_amt = b_row[bank_amt_col]
        candidates = rem_a[rem_a['業務金額'].abs() == b_amt]
        if candidates.empty: continue
        kw_winner = None
        for _, cand in candidates.iterrows():
            km, kw = keyword_match(b_row.get('附言',''), cand.get('描述.1',''))
            if km:
                kw_winner = (cand, kw); break
        if kw_winner:
            a_sel, kw = kw_winner
            print(f"  ✅ [1-A] Bank:{b_row['Bank_index']}({b_amt:,.0f}) ↔ Acc:{a_sel['Acc_index']} 關鍵字「{kw}」")
            history.append({'Type': f'{label} 1對1', 'MatchReason': f'關鍵字優先「{kw}」',
                            'Bank_Data': b_row.to_frame().T, 'Acc_Data': a_sel.to_frame().T})
            rem_a = rem_a[rem_a['Acc_index'] != a_sel['Acc_index']]
            rem_b = rem_b.drop(b_idx)

    print(f"\n⏳ [Step 1-B] 金額唯一輪... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    # 輪二：關鍵字輪結束後，剩餘筆才做金額唯一配對或 LLM
    for b_idx, b_row in rem_b.copy().iterrows():
        if b_idx not in rem_b.index: continue   # 已被同輪 LLM 配走
        b_amt = b_row[bank_amt_col]
        candidates = rem_a[rem_a['業務金額'].abs() == b_amt]
        if len(candidates) == 0:
            continue
        # 同金額的其他銀行筆（可能也在競爭同一筆會計）
        competing_banks = rem_b[(rem_b[bank_amt_col] == b_amt) & (rem_b.index != b_idx)]
        if len(candidates) == 1 and competing_banks.empty:
            # 真正唯一（銀行唯一 + 會計唯一）
            a_row = candidates.iloc[0]
            history.append({'Type': f'{label} 1對1', 'MatchReason': '金額唯一',
                            'Bank_Data': b_row.to_frame().T, 'Acc_Data': a_row.to_frame().T})
            rem_a = rem_a.drop(candidates.index[0])
            rem_b = rem_b.drop(b_idx)
        elif len(candidates) == 1 and not competing_banks.empty:
            # 會計唯一但多筆銀行同金額競爭 → LLM 從銀行端選最吻合的
            a_row = candidates.iloc[0]
            all_b_cands = pd.concat([b_row.to_frame().T, competing_banks])
            print(f"  ⚠️ Acc:{a_row['Acc_index']}({b_amt:,.0f}) 有{len(all_b_cands)}筆銀行競爭，LLM 從銀行端選...")
            best_bank_id = llm_pick_best_bank_candidate(a_row, all_b_cands, mode)
            if best_bank_id:
                b_win = all_b_cands[all_b_cands['Bank_index'] == best_bank_id].iloc[0]
                b_win_idx = all_b_cands[all_b_cands['Bank_index'] == best_bank_id].index[0]
                history.append({'Type': f'{label} 1對1', 'MatchReason': 'LLM確認(銀行端競爭)',
                                'Bank_Data': b_win.to_frame().T, 'Acc_Data': a_row.to_frame().T})
                rem_a = rem_a.drop(candidates.index[0])
                if b_win_idx in rem_b.index:
                    rem_b = rem_b.drop(b_win_idx)
        elif len(candidates) > 1:
            print(f"  ⚠️ Bank:{b_row['Bank_index']}({b_amt:,.0f}) 有{len(candidates)}筆候選，關鍵字無法分辨 → LLM...")
            best_id = llm_pick_best_candidate(b_row, candidates, mode)
            if best_id:
                a_sel = candidates[candidates['Acc_index'] == best_id].iloc[0]
                history.append({'Type': f'{label} 1對1', 'MatchReason': 'LLM確認',
                                'Bank_Data': b_row.to_frame().T, 'Acc_Data': a_sel.to_frame().T})
                rem_a = rem_a[rem_a['Acc_index'] != best_id]
                rem_b = rem_b.drop(b_idx)

    # Step 2: 1 Bank → 2 Acc
    print(f"\n⏳ [Step 2] 1 Bank → 2 Acc... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    for (a1i, a1), (a2i, a2) in itertools.combinations(rem_a.iterrows(), 2):
        if a1['Acc_index'] in matched_a or a2['Acc_index'] in matched_a: continue
        s = abs(a1['業務金額']) + abs(a2['業務金額'])
        for b_idx, b_row in rem_b[rem_b[bank_amt_col] == s].iterrows():
            if b_row['Bank_index'] in matched_b: continue
            is_v, rs = evaluate_combination_with_llm(pd.DataFrame([b_row]), pd.DataFrame([a1, a2]), "1對2", mode)
            if is_v:
                history.append({'Type': f'{label} 1對2', 'MatchReason': rs,
                                'Bank_Data': b_row.to_frame().T,
                                'Acc_Data': pd.concat([a1.to_frame().T, a2.to_frame().T])})
                matched_b.append(b_row['Bank_index']); matched_a.extend([a1['Acc_index'], a2['Acc_index']])
                break  # 此 acc 組合已配對，停止找下一筆銀行
            else:
                print(f"  🟡 1:2 數學吻合但LLM不確定 → 待確認 Bank:{b_row['Bank_index']}")
                pending.append({
                    'Mode': mode, 'Type': '1對2 待確認',
                    'Bank_Data': b_row.to_frame().T,
                    'Acc_Data': pd.concat([a1.to_frame().T, a2.to_frame().T]),
                    'Reason': rs
                })
                break  # 已進入待確認，同樣停止，避免重複

    rem_b = rem_b[~rem_b['Bank_index'].isin(matched_b)]
    rem_a = rem_a[~rem_a['Acc_index'].isin(matched_a)]

    # Step 3: N Bank → 1 Acc
    print(f"\n⏳ [Step 3] N Bank → 1 Acc... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    for a_idx, a_row in rem_a.iterrows():
        if a_row['Acc_index'] in matched_a: continue
        a_amt = abs(a_row['業務金額'])
        if a_amt == 0: continue
        pool = rem_b[(~rem_b['Bank_index'].isin(matched_b)) & (rem_b[bank_amt_col] <= a_amt)]
        found = False
        for r in range(2, min(5, len(pool) + 1)):
            if found: break
            for combo in itertools.combinations(pool.iterrows(), r):
                rows = [c[1] for c in combo]
                if sum(row[bank_amt_col] for row in rows) != a_amt: continue
                combo_df = pd.DataFrame(rows)
                combo_ids = combo_df['Bank_index'].tolist()
                print(f"  💡 N:1 候選 Acc:{a_row['Acc_index']}({a_amt:,.0f}) ← Bank:{combo_ids} → LLM驗證...")
                is_v, rs = evaluate_combination_with_llm(combo_df, a_row.to_frame().T, f"{r}對1", mode)
                if is_v:
                    history.append({'Type': f'{label} {r}對1', 'MatchReason': rs,
                                    'Bank_Data': combo_df, 'Acc_Data': a_row.to_frame().T})
                    matched_b.extend(combo_ids); matched_a.append(a_row['Acc_index'])
                    found = True; break
                else:
                    print(f"  🟡 N:1 數學吻合但LLM不確定 → 待確認 Bank:{combo_ids}")
                    pending.append({
                        'Mode': mode, 'Type': f'{r}對1 待確認',
                        'Bank_Data': combo_df, 'Acc_Data': a_row.to_frame().T, 'Reason': rs
                    })

    rem_b = rem_b[~rem_b['Bank_index'].isin(matched_b)]
    rem_a = rem_a[~rem_a['Acc_index'].isin(matched_a)]

    # Step 4: 同金額 N對N 批次
    print(f"\n⏳ [Step 4] N對N 同金額批次... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    for b_amt, _ in rem_b.groupby(bank_amt_col):
        if rem_b.empty or rem_a.empty or b_amt == 0: break
        b_grp = rem_b[rem_b[bank_amt_col] == b_amt]
        a_grp = rem_a[rem_a['業務金額'].abs() == b_amt]
        if len(b_grp) == 0 or len(a_grp) == 0 or len(b_grp) != len(a_grp): continue
        history.append({'Type': f'{label} {len(b_grp)}對{len(a_grp)}批次', 'Bank_Data': b_grp, 'Acc_Data': a_grp})
        rem_b = rem_b[~rem_b['Bank_index'].isin(b_grp['Bank_index'])]
        rem_a = rem_a[~rem_a['Acc_index'].isin(a_grp['Acc_index'])]

    # Step 5: 同摘要整批加總
    print(f"\n⏳ [Step 5] 同摘要整批加總... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    if '摘要' in rem_b.columns:
        for memo, b_grp in rem_b.groupby('摘要'):
            b_ids = b_grp['Bank_index'].tolist()
            if any(bid in matched_b for bid in b_ids): continue
            b_sum = b_grp[bank_amt_col].sum()
            if b_sum == 0: continue
            cands = rem_a[(rem_a['業務金額'].abs() == b_sum) & (~rem_a['Acc_index'].isin(matched_a))]
            if len(cands) == 1:
                a_row = cands.iloc[0]
                print(f"  ✅ 整批[{memo}] {len(b_grp)}筆合計{b_sum:,.0f} ↔ Acc:{a_row['Acc_index']}")
                history.append({'Type': f'{label} {len(b_grp)}對1整批', 'MatchReason': f'同摘要({memo})整批加總吻合',
                                'Bank_Data': b_grp, 'Acc_Data': a_row.to_frame().T})
                matched_b.extend(b_ids)
                matched_a.append(a_row['Acc_index'])
    rem_b = rem_b[~rem_b['Bank_index'].isin(matched_b)]
    rem_a = rem_a[~rem_a['Acc_index'].isin(matched_a)]

    # Step 6: Coin-change 配對
    print(f"\n⏳ [Step 6] Coin-change配對... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    acc_large = rem_a[rem_a['業務金額'].abs() >= COIN_MIN_AMT].copy()
    acc_large = acc_large.iloc[acc_large['業務金額'].abs().argsort()[::-1]]
    for _, a_row in acc_large.iterrows():
        if a_row['Acc_index'] in matched_a: continue
        a_amt = int(round(abs(a_row['業務金額'])))
        pool  = rem_b[~rem_b['Bank_index'].isin(matched_b)]
        if pool.empty: break
        sel_ids, remainder = coin_fill(pool, bank_amt_col, a_amt)
        if not sel_ids: continue
        is_exact = (remainder == 0)
        is_tol   = (0 < remainder <= COIN_TOLERANCE)
        if is_exact or is_tol:
            reason = 'Coin精確' if is_exact else f'Coin容差(差額{remainder:,}元)'
            sel_df = rem_b[rem_b['Bank_index'].isin(sel_ids)]
            print(f"  ✅ {reason}: {len(sel_ids)}筆合計{sel_df[bank_amt_col].sum():,.0f} → Acc:{a_row['Acc_index']}({a_amt:,.0f})")
            history.append({'Type': f'{label} {len(sel_ids)}對1 Coin', 'MatchReason': reason,
                            'Bank_Data': sel_df, 'Acc_Data': a_row.to_frame().T})
            matched_b.extend(sel_ids)
            matched_a.append(a_row['Acc_index'])
    rem_b = rem_b[~rem_b['Bank_index'].isin(matched_b)]
    rem_a = rem_a[~rem_a['Acc_index'].isin(matched_a)]

    print(f"\n🎉 [{label}] 完成，已核銷 {len(history)} 筆，剩餘銀行{len(rem_b)}/會計{len(rem_a)}")
    return history, rem_b, rem_a


# ── 7. 跨日對帳 ──
import numpy as np

def get_date_str(df, date_col):
    return pd.to_datetime(df[date_col], errors='coerce').dt.strftime('%Y/%m/%d')

def cross_day_reconcile(rem_b, rem_a, bank_date_col, acc_date_col, mode, day_window=7):
    history = []
    rem_b, rem_a = rem_b.copy(), rem_a.copy()
    amt_b = '存入金額' if mode == 'Income' else '支出金額'
    label = '收入' if mode == 'Income' else '支出'
    rem_b['_bdate'] = pd.to_datetime(rem_b[bank_date_col], errors='coerce')
    rem_a['_adate'] = pd.to_datetime(rem_a[acc_date_col],  errors='coerce')
    for b_idx, b_row in rem_b.sort_values('_bdate', ascending=False).iterrows():
        b_amt, b_date = b_row[amt_b], b_row['_bdate']
        cands = rem_a[
            (rem_a['業務金額'].abs() == b_amt) &
            ((rem_a['_adate'] - b_date).abs().dt.days <= day_window)
        ].copy()
        if cands.empty: continue
        cands['_diff'] = (cands['_adate'] - b_date).abs().dt.days
        cands = cands.sort_values('_diff')
        if len(cands) == 1:
            a_row = cands.iloc[0]
            print(f"  🔀 跨日[{label}] Bank:{b_row['Bank_index']}({b_date.date()}) ↔ Acc:{a_row['Acc_index']}({a_row['_adate'].date()}) 差{cands.iloc[0]['_diff']}天")
            history.append({'Type': f'跨日{label}', 'Bank_Data': b_row.to_frame().T, 'Acc_Data': a_row.to_frame().T})
            rem_a = rem_a.drop(cands.index[0]); rem_b = rem_b.drop(b_idx)
        else:
            best_id = llm_pick_best_candidate(b_row, cands, mode)
            if best_id:
                a_sel = cands[cands['Acc_index'] == best_id].iloc[0]
                print(f"  🔀 跨日[{label}][LLM] Bank:{b_row['Bank_index']} ↔ Acc:{best_id}")
                history.append({'Type': f'跨日{label}(LLM)', 'Bank_Data': b_row.to_frame().T, 'Acc_Data': a_sel.to_frame().T})
                rem_a = rem_a[rem_a['Acc_index'] != best_id]; rem_b = rem_b.drop(b_idx)
    rem_b = rem_b.drop(columns=['_bdate'], errors='ignore')
    rem_a = rem_a.drop(columns=['_adate'], errors='ignore')
    return history, rem_b, rem_a

def final_sweep_cross_day(rem_b, rem_a, bank_date_col, acc_date_col, mode, day_window=7, pending=None):
    if pending is None: pending = []
    history = []
    rem_b = rem_b.copy()
    rem_a = rem_a.copy()
    amt_b = '存入金額' if mode == 'Income' else '支出金額'
    label = '收入' if mode == 'Income' else '支出'
    rem_b['_bdate'] = pd.to_datetime(rem_b[bank_date_col], errors='coerce')
    rem_a['_adate'] = pd.to_datetime(rem_a[acc_date_col],  errors='coerce')
    matched_b, matched_a = [], []
    for b_idx, b_row in rem_b.iterrows():
        if b_row['Bank_index'] in matched_b: continue
        b_amt  = b_row[amt_b]
        b_date = b_row['_bdate']
        pool = rem_a[
            (~rem_a['Acc_index'].isin(matched_a)) &
            ((rem_a['_adate'] - b_date).abs().dt.days <= day_window)
        ]
        found = False
        for r in range(2, min(4, len(pool) + 1)):
            if found: break
            for combo in itertools.combinations(pool.iterrows(), r):
                rows = [c[1] for c in combo]
                if sum(abs(row['業務金額']) for row in rows) != b_amt: continue
                combo_df = pd.DataFrame(rows)
                combo_ids = combo_df['Acc_index'].tolist()
                is_v, rs = evaluate_combination_with_llm(b_row.to_frame().T, combo_df, f"跨日1對{r}", mode)
                if is_v:
                    history.append({'Type': f'跨日{label} 1對{r}', 'Bank_Data': b_row.to_frame().T, 'Acc_Data': combo_df})
                    matched_b.append(b_row['Bank_index'])
                    matched_a.extend(combo_ids)
                    found = True; break
                else:
                    print(f"  🟡 跨日1→{r} [{label}] Bank:{b_row['Bank_index']} 數學吻合但LLM不確定 → 待確認")
                    pending.append({
                        'Mode': mode, 'Type': f'跨日{label} 1對{r} 待確認',
                        'Bank_Data': b_row.to_frame().T, 'Acc_Data': combo_df, 'Reason': rs
                    })
    rem_b = rem_b[~rem_b['Bank_index'].isin(matched_b)].drop(columns=['_bdate'], errors='ignore')
    rem_a = rem_a[~rem_a['Acc_index'].isin(matched_a)].drop(columns=['_adate'], errors='ignore')
    return history, rem_b, rem_a


# ── 8. 掃底 ──
def find_soft_matches(rem_b, rem_a, mode):
    soft_matches = []
    amt_col_b = '存入金額' if mode == 'Income' else '支出金額'
    for b_idx, b_row in rem_b.iterrows():
        b_amt = b_row[amt_col_b]
        if b_amt == 0: continue
        candidates = rem_a[rem_a['業務金額'].abs() == b_amt]
        if not candidates.empty:
            soft_matches.append({'Mode': mode, 'Bank_Data': b_row, 'Candidates': candidates})
    return soft_matches


# ── 9. Excel 報表 ──
def build_excel_report(bank_path, inc_rem_b, exp_rem_b, soft_matches, output_path,
                       bank_in, bank_out, acc_in, acc_out, pending_combos=None):
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from copy import copy

    soft_ids = set(m['Bank_Data']['Bank_index'] for m in soft_matches)
    if pending_combos:
        for p in pending_combos:
            for bid in p['Bank_Data']['Bank_index'].values:
                soft_ids.add(str(bid))
    all_rem_b = pd.concat([inc_rem_b, exp_rem_b])
    hard_rem_ids = set(all_rem_b[~all_rem_b['Bank_index'].isin(soft_ids)]['Bank_index'].astype(str))
    pending_ids  = soft_ids

    wb = load_workbook(bank_path)
    ws = wb.active
    print(f'使用工作表：{ws.title}')

    header_row = None
    memo_col = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == '附言':
                header_row = cell.row
                memo_col = cell.column
                break
        if header_row:
            break

    if memo_col is None:
        insert_col = ws.max_column + 1
    else:
        insert_col = memo_col + 1
        ws.insert_cols(insert_col, 2)

    center = Alignment(horizontal='center', vertical='center')
    ref_header = ws.cell(row=header_row, column=memo_col)
    for offset, col_name in enumerate(['未入帳', '待確認']):
        c = ws.cell(row=header_row, column=insert_col + offset, value=col_name)
        c.font      = copy(ref_header.font)
        c.fill      = copy(ref_header.fill)
        c.border    = copy(ref_header.border)
        c.alignment = copy(ref_header.alignment)

    data_start_row = header_row + 1
    for excel_row in range(data_start_row, ws.max_row + 1):
        ref_data = ws.cell(row=excel_row, column=memo_col)
        for col_offset in [0, 1]:
            c = ws.cell(row=excel_row, column=insert_col + col_offset)
            c.fill = copy(ref_data.fill); c.border = copy(ref_data.border)
            c.alignment = center; c.font = copy(ref_data.font)
    for bank_id in hard_rem_ids:
        excel_row = data_start_row + int(bank_id)
        if excel_row <= ws.max_row:
            ws.cell(row=excel_row, column=insert_col).value = 'V'
            ws.cell(row=excel_row, column=insert_col).font  = Font(bold=True, color='FF0000')
    for bank_id in pending_ids:
        excel_row = data_start_row + int(bank_id)
        if excel_row <= ws.max_row:
            ws.cell(row=excel_row, column=insert_col + 1).value = 'V'
            ws.cell(row=excel_row, column=insert_col + 1).font  = Font(bold=True, color='FF8C00')

    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width   = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col + 1)].width = 10

    # 按日期彙總
    summary_row = ws.max_row + 2
    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    center = Alignment(horizontal='center')
    bank_df = pd.concat([bank_in, bank_out])
    acc_df  = pd.concat([acc_in,  acc_out])
    bank_df['_date'] = pd.to_datetime(bank_df['交易日期'], errors='coerce').dt.strftime('%Y/%m/%d')
    acc_df['_date']  = acc_df['業務日期_格式化'] if '業務日期_格式化' in acc_df.columns else pd.to_datetime(acc_df['業務日期'], errors='coerce').dt.strftime('%Y/%m/%d')
    all_dates = sorted(set(bank_df['_date'].dropna()) | set(acc_df['_date'].dropna()))
    col_start = 1
    c = ws.cell(row=summary_row, column=col_start, value='銀行對賬單')
    c.font = Font(bold=True, size=12); c.fill = header_fill
    c = ws.cell(row=summary_row, column=col_start + 4, value='會計賬務')
    c.font = Font(bold=True, size=12); c.fill = PatternFill('solid', fgColor='E2EFDA')
    for i, h in enumerate(['交易日期', '銀行收入', '銀行支出', '銀行變動數']):
        c = ws.cell(row=summary_row+1, column=col_start+i, value=h)
        c.font = bold; c.fill = header_fill; c.alignment = center
    for i, h in enumerate(['交易日期', '會計收入', '會計支出', '會計變動數']):
        c = ws.cell(row=summary_row+1, column=col_start+4+i, value=h)
        c.font = bold; c.fill = PatternFill('solid', fgColor='E2EFDA'); c.alignment = center
    for r, date in enumerate(all_dates):
        row = summary_row + 2 + r
        b_in  = bank_df[(bank_df['_date']==date) & (bank_df['存入金額']>0)]['存入金額'].sum() if '存入金額' in bank_df else 0
        b_out = bank_df[(bank_df['_date']==date) & (bank_df['支出金額']>0)]['支出金額'].sum() if '支出金額' in bank_df else 0
        a_in  = acc_df[(acc_df['_date']==date) & (acc_df['業務金額']>0)]['業務金額'].sum()
        a_out = abs(acc_df[(acc_df['_date']==date) & (acc_df['業務金額']<0)]['業務金額'].sum())
        for col, val in zip(
            [col_start, col_start+1, col_start+2, col_start+3,
             col_start+4, col_start+5, col_start+6, col_start+7],
            [date, b_in, b_out, b_in-b_out, date, a_in, a_out, abs(a_in-a_out)]
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = center
            if isinstance(val, float) or (isinstance(val, int) and col != col_start and col != col_start+4):
                c.number_format = '#,##0'

    wb.save(output_path)
    print(f'\n💾 Excel 報表已產出：{output_path}')


# ── 10. 會計內部對消 ──
ACC_REVERSAL_MAX_COMBO = 4   # 最多幾筆正方合計抵一筆負方（1:1, 2:1, 3:1, 4:1）

def cancel_acc_internal_reversals(inc_rem_a, exp_rem_a):
    """
    純金額對消：從會計端剩餘中找出「N 筆正方合計 = 1 筆負方」的組合，
    將其移出「會計多入」清單。不依賴描述關鍵字，只看金額是否精確抒消。
    支援 1:1、2:1、3:1 … 至 ACC_REVERSAL_MAX_COMBO:1。
    """
    paired_inc, paired_exp = [], []
    log = []

    for _, e_row in exp_rem_a.iterrows():
        if e_row['Acc_index'] in paired_exp: continue
        target = abs(e_row['業務金額'])
        if target == 0: continue

        pool = inc_rem_a[~inc_rem_a['Acc_index'].isin(paired_inc)]

        found = False
        for n in range(1, min(ACC_REVERSAL_MAX_COMBO + 1, len(pool) + 1)):
            if found: break
            for combo in itertools.combinations(pool.iterrows(), n):
                rows = [c[1] for c in combo]
                if sum(abs(r['業務金額']) for r in rows) != target: continue
                combo_ids = [r['Acc_index'] for r in rows]
                log.append({'inc_rows': rows, 'exp': e_row, 'amount': target, 'n': n})
                paired_inc.extend(combo_ids)
                paired_exp.append(e_row['Acc_index'])
                tag = f'{n}:1' if n > 1 else '1:1'
                print(f"  🟣 會計內對消({tag})：{combo_ids}(+{target:,.0f}) ⇔ {e_row['Acc_index']}(-{target:,.0f})")
                found = True
                break

    clean_inc = inc_rem_a[~inc_rem_a['Acc_index'].isin(paired_inc)]
    clean_exp = exp_rem_a[~exp_rem_a['Acc_index'].isin(paired_exp)]
    return clean_inc, clean_exp, log


# ── 10.5 會計迴轉預處理（在銀行配對前先移除沖銷對）──
# '沖回' 故意排除：會出現在「當沖回補」等正常券商業務描述中，導致誤命中
PRE_REVERSAL_KEYWORDS = ['款項更正', '迴轉原傳票', '迴轉', '回轉', '沖銷', '沖正', '轉回']

def preprocess_acc_reversals(acc_in, acc_out):
    """
    銀行配對開始前，先識別「負方描述含迴轉關鍵字且與正方金額完全相符」的對，
    將兩者從配對池移除，避免原始分錄被銀行搶配後迴轉分錄成為孤立多入。
    配對條件：正方描述必須被包含於迴轉描述中（不接受純金額唯一的退而求其次），
    以確保確實是在沖銷哪一張原傳票，而非湊金額的誤配。
    """
    paired_inc, paired_exp = [], []
    log = []

    reversal_entries = acc_out[acc_out['描述.1'].apply(
        lambda d: any(kw in str(d) for kw in PRE_REVERSAL_KEYWORDS)
    )]

    for _, e_row in reversal_entries.iterrows():
        if e_row['Acc_index'] in paired_exp: continue
        target = abs(e_row['業務金額'])
        if target == 0: continue

        pool = acc_in[~acc_in['Acc_index'].isin(paired_inc)]
        candidates = pool[pool['業務金額'].abs() == target]
        if candidates.empty: continue

        # 必須找到「正方描述被包含於迴轉描述」的才配對，不做金額唯一的退而求其次
        # 理由：迴轉傳票命名慣例通常含原傳票全文，若不符則視為不同事件
        e_desc = str(e_row.get('描述.1', ''))
        best = None
        for _, c_row in candidates.iterrows():
            c_desc = str(c_row.get('描述.1', ''))
            if c_desc and len(c_desc) >= 4 and c_desc in e_desc:
                best = c_row; break

        if best is not None:
            log.append({'inc': best, 'exp': e_row, 'amount': target})
            paired_inc.append(best['Acc_index'])
            paired_exp.append(e_row['Acc_index'])
            print(f"  🔵 預處理對消：Acc:{best['Acc_index']}(+{target:,.0f}) ⇔ Acc:{e_row['Acc_index']}(-{target:,.0f})")
            print(f"       原始：{str(best.get('描述.1',''))[:50]}")
            print(f"       迴轉：{str(e_row.get('描述.1',''))[:50]}")

    clean_inc = acc_in[~acc_in['Acc_index'].isin(paired_inc)]
    clean_exp = acc_out[~acc_out['Acc_index'].isin(paired_exp)]
    return clean_inc, clean_exp, log


# ── 11. 淨額配對（迴轉分錄）──
NET_MAX_POS = 2   # 最多幾筆正方合計
NET_MAX_NEG = 2   # 最多幾筆負方（迴轉）合計
NET_REVERSAL_KEYWORDS = ['迴轉', '回轉', '沖銷', '沖正', '轉回', '沖回']

def _is_reversal_entry(acc_row):
    """判斷一筆會計分錄是否為迴轉/沖正性質（描述含迴轉關鍵字）。"""
    desc = str(acc_row.get('描述.1', ''))
    return any(kw in desc for kw in NET_REVERSAL_KEYWORDS)

def net_acc_reconcile(inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a):
    """
    處理「會計有迴轉分錄，導致正方合計 - 負方迴轉 = 銀行金額」的場景。
    例：銀行收 93,339 = 會計 +93,623 + 迴轉 -284
        銀行付 50,000 = 會計 -50,500 + 迴轉 +500
    只在正負方各至少 1 筆的組合才觸發（純同方向的交由既有步驟處理）。
    """
    net_inc_hist, net_exp_hist = [], []
    matched_b_inc, matched_b_exp = [], []
    matched_inc_a, matched_exp_a = [], []
    print(f"\n⏳ [淨額配對] 迴轉分錄搜尋... (銀行收{len(inc_rem_b)}/支{len(exp_rem_b)} | 會計正{len(inc_rem_a)}/負{len(exp_rem_a)})")

    # --- 收款：sum(正方) - sum(迴轉負方) = 銀行存入 ---
    for b_idx, b_row in inc_rem_b.iterrows():
        if b_row['Bank_index'] in matched_b_inc: continue
        b_amt = b_row['存入金額']
        found = False
        avail_pos = inc_rem_a[~inc_rem_a['Acc_index'].isin(matched_inc_a)]
        avail_neg = exp_rem_a[~exp_rem_a['Acc_index'].isin(matched_exp_a)]
        if avail_pos.empty or avail_neg.empty: continue
        for np_ in range(1, min(NET_MAX_POS + 1, len(avail_pos) + 1)):
            if found: break
            for nq_ in range(1, min(NET_MAX_NEG + 1, len(avail_neg) + 1)):
                if found: break
                for pos_combo in itertools.combinations(avail_pos.iterrows(), np_):
                    if found: break
                    pos_rows = [r[1] for r in pos_combo]
                    pos_sum = sum(abs(r['業務金額']) for r in pos_rows)
                    if pos_sum <= b_amt: continue  # 正方合計必須大於銀行金額
                    for neg_combo in itertools.combinations(avail_neg.iterrows(), nq_):
                        neg_rows = [r[1] for r in neg_combo]
                        # 負方至少一筆必須含迴轉關鍵字，避免純數學湊合的假配
                        if not any(_is_reversal_entry(r) for r in neg_rows): continue
                        neg_sum = sum(abs(r['業務金額']) for r in neg_rows)
                        if pos_sum - neg_sum != b_amt: continue
                        pos_df = pd.DataFrame(pos_rows)
                        neg_df = pd.DataFrame(neg_rows)
                        reason = f'正方{pos_sum:,.0f} - 迴轉{neg_sum:,.0f} = 銀行{b_amt:,.0f}'
                        print(f"  ✅ 收款淨額 Bank:{b_row['Bank_index']}({b_amt:,.0f}) = +{pos_sum:,.0f} - {neg_sum:,.0f}")
                        net_inc_hist.append({
                            'Type': f'收入 淨額({np_}正{nq_}負)',
                            'MatchReason': reason,
                            'Bank_Data': b_row.to_frame().T,
                            'Acc_Data': pd.concat([pos_df, neg_df])
                        })
                        matched_b_inc.append(b_row['Bank_index'])
                        matched_inc_a.extend([r['Acc_index'] for r in pos_rows])
                        matched_exp_a.extend([r['Acc_index'] for r in neg_rows])
                        found = True; break

    # --- 付款：sum(|負方|) - sum(迴轉正方) = 銀行支出 ---
    for b_idx, b_row in exp_rem_b.iterrows():
        if b_row['Bank_index'] in matched_b_exp: continue
        b_amt = b_row['支出金額']
        found = False
        avail_neg = exp_rem_a[~exp_rem_a['Acc_index'].isin(matched_exp_a)]
        avail_pos = inc_rem_a[~inc_rem_a['Acc_index'].isin(matched_inc_a)]
        if avail_neg.empty or avail_pos.empty: continue
        for nq_ in range(1, min(NET_MAX_NEG + 1, len(avail_neg) + 1)):
            if found: break
            for np_ in range(1, min(NET_MAX_POS + 1, len(avail_pos) + 1)):
                if found: break
                for neg_combo in itertools.combinations(avail_neg.iterrows(), nq_):
                    if found: break
                    neg_rows = [r[1] for r in neg_combo]
                    neg_sum = sum(abs(r['業務金額']) for r in neg_rows)
                    if neg_sum <= b_amt: continue
                    for pos_combo in itertools.combinations(avail_pos.iterrows(), np_):
                        pos_rows = [r[1] for r in pos_combo]
                        # 正方（迴轉）至少一筆必須含迴轉關鍵字
                        if not any(_is_reversal_entry(r) for r in pos_rows): continue
                        pos_sum = sum(abs(r['業務金額']) for r in pos_rows)
                        if neg_sum - pos_sum != b_amt: continue
                        neg_df = pd.DataFrame(neg_rows)
                        pos_df = pd.DataFrame(pos_rows)
                        reason = f'負方{neg_sum:,.0f} - 迴轉{pos_sum:,.0f} = 銀行{b_amt:,.0f}'
                        print(f"  ✅ 付款淨額 Bank:{b_row['Bank_index']}({b_amt:,.0f}) = -{neg_sum:,.0f} + {pos_sum:,.0f}")
                        net_exp_hist.append({
                            'Type': f'支出 淨額({nq_}負{np_}正)',
                            'MatchReason': reason,
                            'Bank_Data': b_row.to_frame().T,
                            'Acc_Data': pd.concat([neg_df, pos_df])
                        })
                        matched_b_exp.append(b_row['Bank_index'])
                        matched_exp_a.extend([r['Acc_index'] for r in neg_rows])
                        matched_inc_a.extend([r['Acc_index'] for r in pos_rows])
                        found = True; break

    inc_rem_b = inc_rem_b[~inc_rem_b['Bank_index'].isin(matched_b_inc)]
    exp_rem_b = exp_rem_b[~exp_rem_b['Bank_index'].isin(matched_b_exp)]
    inc_rem_a = inc_rem_a[~inc_rem_a['Acc_index'].isin(matched_inc_a)]
    exp_rem_a = exp_rem_a[~exp_rem_a['Acc_index'].isin(matched_exp_a)]
    print(f"  淨額配對完成：已配 收入{len(net_inc_hist)}/支出{len(net_exp_hist)} 組")
    return net_inc_hist, net_exp_hist, inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a


# ── 12. HTML 報告 ──
def build_html_report_with_soft_match(all_histories, all_rems, soft_matches, date, output_path,
                                      pending_combos=None, reversal_log=None,
                                      acc_only_rems=None, acc_reversal_log=None,
                                      pre_reversal_log=None):
    style = """
    <style>
        body { font-family: 'Inter', system-ui, sans-serif; background: #f8fafc; color: #1e293b; padding: 40px; }
        .container { max-width: 1400px; margin: auto; }
        h1 { color: #0f172a; border-bottom: 3px solid #3b82f6; padding-bottom: 10px; }
        .summary-box { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 30px; }
        .stat-card { background: white; padding: 20px; border-radius: 12px; box-shadow: 0 4px 6px rgb(0 0 0 / 0.05); }
        .card { background: white; border-radius: 16px; padding: 25px; margin-bottom: 40px; box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1); }
        .tag { display: inline-block; padding: 4px 12px; border-radius: 999px; font-size: 12px; font-weight: bold; margin-bottom: 15px; }
        .tag-inc { background: #dcfce7; color: #166534; }
        .tag-exp { background: #fee2e2; color: #991b1b; }
        .tag-warn { background: #fef08a; color: #854d0e; }
        .table-section { margin-top: 15px; overflow-x: auto; }
        .side-title { font-size: 13px; text-transform: uppercase; font-weight: bold; color: #64748b; margin-bottom: 10px; border-left: 4px solid #3b82f6; padding-left: 10px; }
        table { width: 100%; border-collapse: collapse; font-size: 12px; margin-bottom: 10px; text-align: right; }
        th { background: #f8fafc; padding: 10px; border-bottom: 2px solid #e2e8f0; text-align: right; }
        td { padding: 10px; border-bottom: 1px solid #f1f5f9; }
    </style>
    """
    total_matched = len(all_histories['Income']) + len(all_histories['Expense'])
    total_pending_all = len(soft_matches) + len(pending_combos or [])
    rem_b_ids_in_soft = [m['Bank_Data']['Bank_index'] for m in soft_matches]
    total_hard_rem = len(all_rems['Income_B'][~all_rems['Income_B']['Bank_index'].isin(rem_b_ids_in_soft)]) + \
                     len(all_rems['Expense_B'][~all_rems['Expense_B']['Bank_index'].isin(rem_b_ids_in_soft)])

    html = f"<div class='container'><h1>收支核銷報告 ({date})</h1>"

    html += "<div class='summary-box'>"
    html += f"<div class='stat-card' style='border-top: 4px solid #22c55e;'><h3>🟢 總結成功核銷</h3><div style='font-size:32px; font-weight:bold; color:#22c55e;'>{total_matched} 筆</div></div>"
    html += f"<div class='stat-card' style='border-top: 4px solid #eab308;'><h3>🟡 全部待確認</h3><div style='font-size:32px; font-weight:bold; color:#eab308;'>{total_pending_all} 筆</div></div>"
    html += f"<div class='stat-card' style='border-top: 4px solid #ef4444;'><h3>🔴 總計剩餘未入帳 (銀行端)</h3><div style='font-size:32px; font-weight:bold; color:#ef4444;'>{total_hard_rem} 筆</div></div>"
    html += "</div>"

    if soft_matches or pending_combos:
        total_p = len(soft_matches) + len(pending_combos or [])
        html += f"<h2>🟡 全部待確認（共 {total_p} 筆）</h2>"
        html += "<p style='color: #64748b; font-size: 14px;'>以下為金額吻合但需人工確認的紀錄，包含掃底配對及組合配對。</p>"
        for idx, match in enumerate(soft_matches):
            mode_tw = '收款' if match['Mode'] == 'Income' else '付款'
            b_data = match['Bank_Data']
            html += f"<div class='card' style='border-left: 5px solid #eab308; background-color: #fefce8; padding: 20px;'>"
            html += f"<div class='tag tag-warn'>待確認 {mode_tw}（1對1金額相符）</div>"
            html += "<div class='table-section'><div class='side-title'>【銀行端】</div>"
            html += b_data.to_frame().T.to_html(index=False)
            html += "</div>"
            html += f"<div class='table-section' style='margin-top:20px;'><div class='side-title'>👇 會計端候選（{len(match['Candidates'])} 筆）</div>"
            html += match['Candidates'].to_html(index=False)
            html += "</div></div>"
        for p in (pending_combos or []):
            html += f"<div class='card' style='border-left:5px solid #eab308;background:#fefce8;padding:20px;'>"
            html += f"<div class='tag tag-warn'>{p['Type']}</div>"
            html += f"<div style='color:#92400e;margin-bottom:10px;'>LLM意見：{p['Reason']}</div>"
            html += "<div class='table-section'><div class='side-title'>【銀行端】</div>"
            html += p['Bank_Data'].to_html(index=False)
            html += "</div><div class='table-section' style='margin-top:15px;'><div class='side-title'>【會計端】</div>"
            html += p['Acc_Data'].to_html(index=False)
            html += "</div></div>"

    html += "<h2 style='margin-top:40px'>🔴 剩餘未入帳之銀行賬務清單 (待盤點)</h2>"
    p_map = {'Income': '尚未入帳之【收款】', 'Expense': '尚未入帳之【付款】'}
    for m in ['Income', 'Expense']:
        df_rem = all_rems[f'{m}_B']
        df_rem_pure = df_rem[~df_rem['Bank_index'].isin(rem_b_ids_in_soft)]
        if not df_rem_pure.empty:
            html += f"<div class='card' style='border-top: 5px solid {'#22c55e' if m=='Income' else '#ef4444'}'>"
            html += f"<div class='tag {'tag-inc' if m=='Income' else 'tag-exp'}'>{p_map[m]}</div>"
            html += df_rem_pure[['Bank_index', '交易日期', '交易時間', '附言', '摘要', '存入金額', '支出金額']].to_html(index=False)
            html += "</div>"

    acc_only_rems = acc_only_rems or {}
    acc_only_items = [
        ('Income',  '【收款】會計多入（銀行無此筆）', '#f97316', 'tag-inc'),
        ('Expense', '【付款】會計多入（銀行無此筆）', '#f97316', 'tag-exp'),
    ]
    has_acc_only = any(not acc_only_rems.get(f'{m}_A', pd.DataFrame()).empty for m, *_ in acc_only_items)
    if has_acc_only:
        html += "<h2 style='margin-top:40px'>🟠 會計多入清單（會計有、銀行無，待查）</h2>"
        html += "<p style='color:#64748b;font-size:14px;'>以下為會計帳內存在、但銀行對帳單找不到對應紀錄的項目，請逐筆確認。</p>"
        acc_cols = ['Acc_index', '業務日期_格式化', '憑證編號', '描述.1', '業務金額', '借方/貸方', '來源Sheet']
        for m, label_txt, border_color, tag_cls in acc_only_items:
            df_ao = acc_only_rems.get(f'{m}_A', pd.DataFrame())
            if df_ao.empty: continue
            show_cols = [c for c in acc_cols if c in df_ao.columns]
            html += f"<div class='card' style='border-top:5px solid {border_color}'>"
            html += f"<div class='tag {tag_cls}'>{label_txt}</div>"
            html += df_ao[show_cols].to_html(index=False)
            html += "</div>"

    if acc_reversal_log:
        html += f"<h2 style='margin-top:40px'>🟣 會計內部對消（共 {len(acc_reversal_log)} 組，已自動移出多入清單）</h2>"
        html += "<p style='color:#64748b;font-size:14px;'>以下為會計帳內金額精確抵消的組合（純金額判斷），無需銀行對應紀錄。</p>"
        acc_cols_show = ['Acc_index', '業務日期_格式化', '憑證編號', '描述.1', '業務金額', '借方/貸方', '來源Sheet']
        for rv in acc_reversal_log:
            amt  = rv['amount']
            n    = rv['n']
            inc_df = pd.DataFrame(rv['inc_rows'])
            exp_df = pd.DataFrame([rv['exp']])
            inc_cols = [c for c in acc_cols_show if c in inc_df.columns]
            exp_cols = [c for c in acc_cols_show if c in exp_df.columns]
            tag_txt  = f"{n}:1 對消配對" if n > 1 else "1:1 對消配對"
            html += "<div class='card' style='border-left:5px solid #7c3aed;background:#faf5ff;padding:20px;'>"
            html += f"<div class='tag' style='background:#ede9fe;color:#4c1d95;'>{tag_txt}  {amt:,.0f} 元</div>"
            html += "<div style='display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:10px;'>"
            html += f"<div><div class='side-title'>正方（{n} 筆，合計 +{amt:,.0f}）</div>"
            html += inc_df[inc_cols].to_html(index=False)
            html += "</div><div><div class='side-title'>負方（1 筆，-{:,.0f}）</div>".format(amt)
            html += exp_df[exp_cols].to_html(index=False)
            html += "</div></div></div>"

    if pre_reversal_log:
        acc_cols_pre = ['Acc_index', '業務日期_格式化', '憑證編號', '描述.1', '業務金額', '借方/貸方', '來源Sheet']
        html += f"<h2 style='margin-top:40px'>🔵 會計迴轉預處理對消（共 {len(pre_reversal_log)} 對，已於配對前移除）</h2>"
        html += "<p style='color:#64748b;font-size:14px;'>以下為「負方含迴轉關鍵字且金額與正方吻合」的對，已在銀行配對前預先對消，避免原始分錄被銀行搶配。對應的銀行收款將顯示為未入帳。</p>"
        for rv in pre_reversal_log:
            inc_row = rv['inc']
            exp_row = rv['exp']
            amt = rv['amount']
            inc_df = pd.DataFrame([inc_row])
            exp_df = pd.DataFrame([exp_row])
            inc_show = [c for c in acc_cols_pre if c in inc_df.columns]
            exp_show = [c for c in acc_cols_pre if c in exp_df.columns]
            html += "<div class='card' style='border-left:5px solid #2563eb;background:#eff6ff;padding:20px;'>"
            html += f"<div class='tag' style='background:#dbeafe;color:#1e40af;'>預處理對消  {amt:,.0f} 元</div>"
            html += "<div style='display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:10px;'>"
            html += f"<div><div class='side-title'>原始分錄（正方 +{amt:,.0f}）</div>"
            html += inc_df[inc_show].to_html(index=False)
            html += f"</div><div><div class='side-title'>迴轉分錄（負方 -{amt:,.0f}）</div>"
            html += exp_df[exp_show].to_html(index=False)
            html += "</div></div></div>"

    if reversal_log:
        html += f"<h2 style='margin-top:40px'>🔄 日內沖銷紀錄（共 {len(reversal_log)} 組）</h2>"
        html += "<p style='color:#64748b;font-size:14px;'>以下轉支與沖正轉支同日金額相符，已於對帳前自動移除。</p>"
        for rv in reversal_log:
            orig, chong, amt = rv['orig'], rv['chong'], rv['amount']
            html += "<div class='card' style='border-left:5px solid #6366f1;background:#f5f3ff;padding:20px;'>"
            html += f"<div class='tag' style='background:#ede9fe;color:#4c1d95;'>沖銷配對  {amt:,.0f} 元</div>"
            html += "<div style='display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:10px;'>"
            html += f"<div><div class='side-title'>原始轉支（Bank_index {orig['Bank_index']}）</div>"
            html += f"<table><tr><th>交易日期</th><th>摘要</th><th>支出金額</th></tr>"
            html += f"<tr><td>{orig.get('交易日期','')}</td><td>{orig.get('摘要','')}</td><td style='color:#dc2626'>{amt:,.0f}</td></tr></table></div>"
            html += f"<div><div class='side-title'>沖正轉支（Bank_index {chong['Bank_index']}）</div>"
            html += f"<table><tr><th>交易日期</th><th>摘要</th><th>支出金額</th></tr>"
            html += f"<tr><td>{chong.get('交易日期','')}</td><td>{chong.get('摘要','')}</td><td style='color:#16a34a'>-{amt:,.0f}</td></tr></table></div>"
            html += "</div></div>"

    html += "<h2 style='margin-top:60px'>🟢 詳盡對帳軌跡 (已核銷)</h2>"
    all_matched = all_histories['Income'] + all_histories['Expense']
    for idx, item in enumerate(all_matched):
        is_inc = '收入' in item['Type']
        reason = item.get('MatchReason', '—')
        html += "<div class='card'>"
        html += f"<div class='tag {'tag-inc' if is_inc else 'tag-exp'}'>{item['Type']}</div>"
        html += f"<div style='font-size:12px;color:#64748b;margin-bottom:10px;'>⚙️ 配對依據：{reason}</div>"
        html += "<div class='table-section'><div class='side-title'>銀行端原始紀錄</div>"
        html += item['Bank_Data'].to_html(index=False)
        html += "</div>"
        html += "<div class='table-section'><div class='side-title'>會計端原始紀錄</div>"
        html += item['Acc_Data'].to_html(index=False)
        html += "</div></div>"

    html += "</div>"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"<html><head><meta charset='utf-8'>{style}</head><body>{html}</body></html>")


# ══════════════════════════════════════════════════
# 12. 單月對帳完整流程
# ══════════════════════════════════════════════════
def run_month_pipeline(month_code):
    target_date = month_code_to_date(month_code)
    bank_path = convert_to_xlsx(bank_files[month_code])
    print(f'\n{"="*60}')
    print(f'🗓️  處理月份：{target_date}  ({os.path.basename(bank_path)})')
    print(f'{"="*60}')

    bank_in, bank_out, acc_in, acc_out, reversal_log = load_and_preprocess(
        target_date, bank_path_param=bank_path, acc_sheets_param=_acc_sheets_global
    )
    print(f"✅ 數據載入：[收入組] 銀行 {len(bank_in)} 筆 / 會計 {len(acc_in)} 筆 | [支出組] 銀行 {len(bank_out)} 筆 / 會計 {len(acc_out)} 筆")
    _total_bank = len(bank_in) + len(bank_out)
    _total_acc  = len(acc_in)  + len(acc_out)
    print(f"   ▶ 銀行合計 {_total_bank} 筆（另有日內沖銷 {len(reversal_log)} 組 = {len(reversal_log)*2} 筆已移除）")
    print(f"   ▶ 會計合計 {_total_acc} 筆（正方 {len(acc_in)} 筆 + 負方 {len(acc_out)} 筆）")

    # 保存原始筆數供稽核用（預處理後 acc_in/acc_out 可能減少）
    _acc_in_orig_count  = len(acc_in)
    _acc_out_orig_count = len(acc_out)

    # ── 預處理：會計迴轉對消（在銀行配對前先移除） ──
    print('\n========== 🔵 會計迴轉預處理 ==========')
    acc_in, acc_out, pre_reversal_log = preprocess_acc_reversals(acc_in, acc_out)
    print(f'預處理完成：移除 {len(pre_reversal_log)} 對迴轉分錄（正方 {len(pre_reversal_log)} 筆 / 負方 {len(pre_reversal_log)} 筆）')

    # ── 階段一：逐日對帳 ──
    bank_dates = sorted(set(
        get_date_str(bank_in, '交易日期').dropna().tolist() +
        get_date_str(bank_out, '交易日期').dropna().tolist()
    ))
    all_inc_hist, all_exp_hist = [], []
    all_inc_rem_b, all_exp_rem_b = [], []
    all_inc_rem_a, all_exp_rem_a = [], []
    all_pending_combos = []

    for date in bank_dates:
        b_in_day  = bank_in[get_date_str(bank_in,   '交易日期') == date]
        b_out_day = bank_out[get_date_str(bank_out, '交易日期') == date]
        a_in_day  = acc_in[acc_in['業務日期_格式化']  == date]
        a_out_day = acc_out[acc_out['業務日期_格式化'] == date]
        print(f'\n===== 📅 {date} ===== 銀行收{len(b_in_day)}/支{len(b_out_day)} | 會計收{len(a_in_day)}/支{len(a_out_day)}')
        h_i, r_bi, r_ai = reconcile_engine(b_in_day,  a_in_day,  'Income',  all_pending_combos)
        h_e, r_be, r_ae = reconcile_engine(b_out_day, a_out_day, 'Expense', all_pending_combos)
        all_inc_hist += h_i;  all_exp_hist += h_e
        all_inc_rem_b.append(r_bi); all_exp_rem_b.append(r_be)
        all_inc_rem_a.append(r_ai); all_exp_rem_a.append(r_ae)

    inc_hist  = all_inc_hist
    exp_hist  = all_exp_hist
    inc_rem_b = pd.concat(all_inc_rem_b) if all_inc_rem_b else pd.DataFrame()
    exp_rem_b = pd.concat(all_exp_rem_b) if all_exp_rem_b else pd.DataFrame()
    inc_rem_a = pd.concat(all_inc_rem_a) if all_inc_rem_a else pd.DataFrame()
    exp_rem_a = pd.concat(all_exp_rem_a) if all_exp_rem_a else pd.DataFrame()

    # ── 階段 1.5：淨額配對（迴轉分錄，需跨正負方會計池） ──
    print('\n\n========== 🔢 淨額配對（階段 1.5） ==========')
    net_inc_h, net_exp_h, inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a = net_acc_reconcile(
        inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a
    )
    inc_hist += net_inc_h
    exp_hist += net_exp_h

    # ── 階段二：跨日對帳 ──
    print('\n\n========== 🔀 跨日對帳（階段二） ==========')
    cross_inc_hist, inc_rem_b, inc_rem_a = cross_day_reconcile(inc_rem_b, inc_rem_a, '交易日期', '業務日期_格式化', 'Income')
    cross_exp_hist, exp_rem_b, exp_rem_a = cross_day_reconcile(exp_rem_b, exp_rem_a, '交易日期', '業務日期_格式化', 'Expense')
    inc_hist += cross_inc_hist
    exp_hist += cross_exp_hist

    sweep_inc_hist, inc_rem_b, inc_rem_a = final_sweep_cross_day(inc_rem_b, inc_rem_a, '交易日期', '業務日期_格式化', 'Income',  pending=all_pending_combos)
    sweep_exp_hist, exp_rem_b, exp_rem_a = final_sweep_cross_day(exp_rem_b, exp_rem_a, '交易日期', '業務日期_格式化', 'Expense', pending=all_pending_combos)
    inc_hist += sweep_inc_hist
    exp_hist += sweep_exp_hist

    print(f'\n{"="*50}')
    print(f'最終剩餘：收入(銀行{len(inc_rem_b)}/會計{len(inc_rem_a)}) | 支出(銀行{len(exp_rem_b)}/會計{len(exp_rem_a)})')
    if not exp_rem_b.empty:
        print(f'\n付款剩餘明細：')
        print(exp_rem_b[['Bank_index', '交易日期', '附言', '支出金額']].to_string())
    print(f"\n🟡 組合配對待確認：{len(all_pending_combos)} 筆")
    for p in all_pending_combos:
        print(f"  [{p['Type']}] {p['Reason']}")
        print(f"  銀行：\n{p['Bank_Data'][['Bank_index','支出金額' if p['Mode']=='Expense' else '存入金額','附言']].to_string()}")
        print(f"  會計：\n{p['Acc_Data'][['Acc_index','業務金額','描述.1' if '描述.1' in p['Acc_Data'].columns else '描述']].to_string()}")
        print()

    # ── 掃底 ──
    inc_soft_matches = find_soft_matches(inc_rem_b, inc_rem_a, 'Income')
    exp_soft_matches = find_soft_matches(exp_rem_b, exp_rem_a, 'Expense')
    all_soft_matches = inc_soft_matches + exp_soft_matches

    # ── Excel 報表 ──
    output_xlsx = os.path.join(data_dir, f'未入帳清單_{month_code}.xlsx')
    build_excel_report(
        bank_path, inc_rem_b, exp_rem_b,
        all_soft_matches, output_xlsx,
        bank_in, bank_out, acc_in, acc_out,
        pending_combos=all_pending_combos
    )

    # ── 會計內部對消 + HTML 報表 ──
    inc_hist = [item for item in inc_hist if isinstance(item, dict)]
    exp_hist = [item for item in exp_hist if isinstance(item, dict)]
    print(f"清理後：inc_hist {len(inc_hist)} 筆 / exp_hist {len(exp_hist)} 筆")

    print('\n\n========== 🟣 會計端內部對消 ==========')
    inc_rem_a, exp_rem_a, acc_reversal_log = cancel_acc_internal_reversals(inc_rem_a, exp_rem_a)
    print(f'對消後剩餘：會計收入 {len(inc_rem_a)} 筆 / 會計支出 {len(exp_rem_a)} 筆')

    # ── 對帳完整性稽核 ──
    print('' + '─'*62)
    print('📋  對帳完整性稽核')
    print('─'*62)
    _m_inc_b  = sum(len(h['Bank_Data']) for h in inc_hist)
    _m_exp_b  = sum(len(h['Bank_Data']) for h in exp_hist)
    _m_inc_a  = sum(len(h['Acc_Data'])  for h in inc_hist)
    _m_exp_a  = sum(len(h['Acc_Data'])  for h in exp_hist)
    _rv_inc   = sum(rv['n'] for rv in acc_reversal_log)
    _rv_exp   = len(acc_reversal_log)
    _pre_rv   = len(pre_reversal_log)   # 預處理移除的對數（正負各 _pre_rv 筆）
    _ok_inc_b = (_m_inc_b + len(inc_rem_b) == len(bank_in))
    _ok_exp_b = (_m_exp_b + len(exp_rem_b) == len(bank_out))
    _ok_inc_a = (_m_inc_a + _rv_inc + _pre_rv + len(inc_rem_a) == _acc_in_orig_count)
    _ok_exp_a = (_m_exp_a + _rv_exp + _pre_rv + len(exp_rem_a) == _acc_out_orig_count)
    print(f"銀行 收入: 原始{len(bank_in):>5} = 核銷{_m_inc_b:>5} + 剩餘{len(inc_rem_b):>5}  {'✅' if _ok_inc_b else '❌ 不平衡！'}")
    print(f"銀行 支出: 原始{len(bank_out):>5} = 核銷{_m_exp_b:>5} + 剩餘{len(exp_rem_b):>5}  {'✅' if _ok_exp_b else '❌ 不平衡！'}")
    print(f"會計 正方: 原始{_acc_in_orig_count:>5} = 核銷{_m_inc_a:>5} + 預沖{_pre_rv:>5} + 對消{_rv_inc:>5} + 多入{len(inc_rem_a):>5}  {'✅' if _ok_inc_a else '❌ 不平衡！'}")
    print(f"會計 負方: 原始{_acc_out_orig_count:>5} = 核銷{_m_exp_a:>5} + 預沖{_pre_rv:>5} + 對消{_rv_exp:>5} + 多入{len(exp_rem_a):>5}  {'✅' if _ok_exp_a else '❌ 不平衡！'}")
    if all([_ok_inc_b, _ok_exp_b, _ok_inc_a, _ok_exp_a]):
        print('✅  每筆均已歸類，分類總和與原始筆數完全吻合。')
    else:
        print('⚠️   有分類不平衡，對帳邏輯可能存在漏入或重複！')
    print('─'*62)

    output_html = os.path.join(data_dir, f'未入帳清單_{month_code}.html')
    build_html_report_with_soft_match(
        {'Income': inc_hist, 'Expense': exp_hist},
        {'Income_B': inc_rem_b, 'Expense_B': exp_rem_b},
        all_soft_matches,
        target_date,
        output_html,
        pending_combos=all_pending_combos,
        reversal_log=reversal_log,
        acc_only_rems={'Income_A': inc_rem_a, 'Expense_A': exp_rem_a},
        acc_reversal_log=acc_reversal_log,
        pre_reversal_log=pre_reversal_log
    )
    print(f'\n✅ [{month_code}] {target_date} 完成！')
    print(f'   HTML  → {os.path.basename(output_html)}')
    print(f'   Excel → {os.path.basename(output_xlsx)}')


# ══════════════════════════════════════════════════
# 13. 主執行迴圈：依月份碼排序，逐月處理
# ══════════════════════════════════════════════════
for month_code in sorted(bank_files.keys()):
    run_month_pipeline(month_code)
