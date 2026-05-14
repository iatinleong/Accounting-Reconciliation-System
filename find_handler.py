"""
find_handler.py
針對未入帳清單（Excel）中每一筆未入帳/待確認的銀行賬務，
找出會計經辦人，並產出 HTML 報告。

資料流：
  1. 掃描 帳務查詢-*.xlsx + 未銷帳明細表_.xlsm → 更新 handler_history.json
  2. 查找順序：未銷帳明細（即時）→ 歷史資料庫 → 固定規則 → 推論 → 人工處理
  3. 輸出：更新未入帳清單_.xlsx（新增「經辦人」欄）+ 未入帳_經辦人報告.html
  所有 input/output 均使用 DATA_DIR（資料\）。
"""

import os, re, glob, difflib, sys, json
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment
from copy import copy
from datetime import datetime

sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1, closefd=False)

# ══════════════════════════════════════════════════
# 0. 設定
# ══════════════════════════════════════════════════

_BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.path.join(_BASE_DIR, '資料')
def _find_ar_path(data_dir: str) -> str:
    for fname in os.listdir(data_dir):
        if '未銷帳明細' in fname and fname.endswith('.xlsm') and not fname.startswith('~$'):
            return os.path.join(data_dir, fname)
    return os.path.join(data_dir, '未銷帳明細表_.xlsm')  # fallback

AR_PATH      = _find_ar_path(DATA_DIR)
HISTORY_PATH = os.path.join(_BASE_DIR, 'handler_history.json')
OUTPUT_DIR   = DATA_DIR

STAFF_TABLE = {
    'A01': '李志洪', 'A02': '廖倖伶', 'A03': '賀芳華', 'A04': '曾玉英',
    'A05': '張敏菁', 'A06': '賴慈香', 'A07': '吳美芬', 'A08': '楊淑娟',
    'A09': '彭雅芬', 'A10': '陳振裕', 'A11': '陳美珍', 'A12': '許嘉芳',
    'A13': '高毓玲', 'A14': '林素雲', 'A15': '吳淑愛', 'A16': '陳宜靖',
    'A17': '陳曉青', 'A18': '李芳姿', 'A19': '吳佩蓉', 'A20': '張鈴宜',
    'A21': '陳雅玲', 'A22': '沈淑慧', 'A23': '陳玉玲', 'A24': '蘇怡珊',
    'A25': '陳彥如', 'A26': '陳秋安', 'A27': '李宜蓁', 'A28': '陳淑瑜',
}

FUZZY_THRESHOLD  = 0.45  # pure fuzzy 路徑的最低分
EXACT_MIN_FUZZY  = 0.15  # same-handler shortcut 的最低相似度（低於此視為文字完全不相關）
MIN_KEYWORD_LEN = 3    # 關鍵字至少需有幾個 CJK 字才算「命中」（2字如「富邦」太泛，不算）
MIN_MATCH_CHARS = 2    # cjk_substrings 產生子字串的最短長度（內部用）

ACC_SKIP = {'tmpscrapsheet', '餘額', '餘額_日',
            'QAA_Internal_WorkSheet', 'QAA_Sample_Data_WorkSheet',
            'Report_Design_Format_HSheet'}
AR_SKIP  = {'QAA_Internal_WorkSheet', 'QAA_Sample_Data_WorkSheet',
            'Report_Design_Format_HSheet'}

def get_tolerance(amount: float) -> float:
    return 500 if amount <= 10_000 else 2_000

# ══════════════════════════════════════════════════
# 1. 歷史資料庫維護
# ══════════════════════════════════════════════════

def _load_history() -> dict:
    if os.path.exists(HISTORY_PATH):
        with open(HISTORY_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def _save_history(history: dict):
    with open(HISTORY_PATH, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def _entry_key(src_file: str, src_sheet: str, desc: str, amount: float) -> str:
    return f"{src_file}|{src_sheet}|{desc}|{amount}"

def update_handler_history() -> dict:
    """
    掃描 DATA_DIR 下所有 帳務查詢-*.xlsx 和 未銷帳明細表_.xlsm，
    將新記錄 append 至 handler_history.json 並存檔。
    已存在的記錄（以 src_file+src_sheet+desc+amount 去重）不重複新增。
    回傳更新後的 history dict。

    history 結構：
    {
      "A74": {
        "name": "羅翊瑄",
        "entries": [
          {
            "src_file": "帳務查詢-11501-11503.xlsx",
            "src_sheet": "11501",
            "src_type": "acc",          # acc=帳務查詢 / ar=未銷帳明細
            "date": "2026-01-15",
            "desc": "114/11~12月中國A股基金收入-華南永昌",
            "amount": 61.0
          },
          ...
        ]
      },
      ...
    }
    """
    history  = _load_history()
    seen     = set()
    for code, data in history.items():
        for e in data.get('entries', []):
            seen.add(_entry_key(e['src_file'], e['src_sheet'], e['desc'], e['amount']))

    new_count = 0

    # ── 掃描帳務查詢-*.xlsx ───────────────────────
    acc_files = sorted(glob.glob(os.path.join(DATA_DIR, '帳務查詢*.xlsx')))
    acc_files = [f for f in acc_files if not os.path.basename(f).startswith('~$')]
    print(f"  掃描帳務查詢：{len(acc_files)} 個檔案")

    for fpath in acc_files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"    無法開啟 {fname}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            if sheet_name in ACC_SKIP:
                continue
            ws = wb[sheet_name]
            sheet_new = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or len(row) < 12:
                    continue
                handler  = str(row[0] or '').strip()
                desc     = str(row[9] or '').strip()
                raw_amt  = row[11]
                name     = str(row[41] or '').strip() if len(row) > 41 else ''
                date_val = row[6] if len(row) > 6 else None
                if not handler or not desc:
                    continue
                try:
                    amt = abs(float(raw_amt or 0))
                except Exception:
                    continue
                if amt <= 0:
                    continue
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_val or '').split(' ')[0]

                key = _entry_key(fname, sheet_name, desc, amt)
                if key in seen:
                    continue
                seen.add(key)
                new_count  += 1
                sheet_new  += 1

                if handler not in history:
                    history[handler] = {
                        'name': name or STAFF_TABLE.get(handler, ''),
                        'entries': []
                    }
                elif name and not history[handler].get('name'):
                    history[handler]['name'] = name

                history[handler]['entries'].append({
                    'src_file':  fname,
                    'src_sheet': sheet_name,
                    'src_type':  'acc',
                    'date':      date_str,
                    'desc':      desc,
                    'amount':    amt,
                })

            if sheet_new:
                print(f"    {fname} / {sheet_name}：新增 {sheet_new} 筆")
        wb.close()

    # ── 掃描未銷帳明細表_.xlsm ───────────────────
    if os.path.exists(AR_PATH):
        fname_ar = os.path.basename(AR_PATH)
        print(f"  掃描未銷帳明細：{fname_ar}")
        try:
            xl = pd.ExcelFile(AR_PATH, engine='openpyxl')
            for sheet_name in xl.sheet_names:
                if sheet_name in AR_SKIP:
                    continue
                df = pd.read_excel(AR_PATH, sheet_name=sheet_name, header=None,
                                   engine='openpyxl', keep_default_na=False)
                if len(df) < 7:
                    continue
                sheet_new = 0
                for _, row in df.iloc[6:].iterrows():
                    amt_raw = row[3] if row[3] != '' else None
                    if amt_raw is None:
                        continue
                    try:
                        amt = abs(float(str(amt_raw).replace(',', '')))
                    except Exception:
                        continue
                    if amt <= 0:
                        continue
                    desc    = str(row[8]).strip()  if row[8]  != '' else ''
                    handler = str(row[18]).strip() if row[18] != '' else ''
                    if not handler or not desc:
                        continue
                    key = _entry_key(fname_ar, sheet_name, desc, amt)
                    if key in seen:
                        continue
                    seen.add(key)
                    new_count += 1
                    sheet_new += 1
                    if handler not in history:
                        history[handler] = {
                            'name': STAFF_TABLE.get(handler, ''),
                            'entries': []
                        }
                    history[handler]['entries'].append({
                        'src_file':  fname_ar,
                        'src_sheet': sheet_name,
                        'src_type':  'ar',
                        'date':      '',
                        'desc':      desc,
                        'amount':    amt,
                    })
                if sheet_new:
                    print(f"    {fname_ar} / {sheet_name}：新增 {sheet_new} 筆")
        except Exception as e:
            print(f"    讀取未銷帳明細失敗: {e}")

    _save_history(history)
    total_entries = sum(len(v['entries']) for v in history.values())
    print(f"  ✅ 歷史資料庫：新增 {new_count} 筆，共 {total_entries} 筆記錄（{len(history)} 位經辦）")
    print(f"     儲存至：{HISTORY_PATH}")
    return history

# ══════════════════════════════════════════════════
# 2. 載入查找用表
# ══════════════════════════════════════════════════

def parse_amount_str(v) -> float:
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0

def load_ar_table(path=AR_PATH) -> pd.DataFrame:
    """載入未銷帳明細表（即時，供方法 1A/3A 使用）"""
    xl   = pd.ExcelFile(path, engine='openpyxl')
    rows = []
    for sheet in xl.sheet_names:
        if sheet in AR_SKIP:
            continue
        df = pd.read_excel(path, sheet_name=sheet, header=None,
                           engine='openpyxl', keep_default_na=False)
        if len(df) < 7:
            continue
        for _, row in df.iloc[6:].iterrows():
            amt_val = row[3] if row[3] != '' else None
            if amt_val is None:
                continue
            amt = parse_amount_str(amt_val)
            if amt <= 0:
                continue
            desc    = str(row[8]).strip()  if row[8]  != '' else ''
            handler = str(row[18]).strip() if row[18] != '' else ''
            rows.append({
                '_handler': handler,
                '_name':    STAFF_TABLE.get(handler, ''),
                '_desc':    desc,
                '_amount':  amt,
                '_date':    '',
                '_file':    os.path.basename(AR_PATH),
                '_sheet':   sheet,
                '_src_type': 'ar',
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        '_handler','_name','_desc','_amount','_date','_file','_sheet','_src_type'])

def history_to_df(history: dict) -> pd.DataFrame:
    """把 history dict 展開成可搜尋的 DataFrame（供方法 1B/3B 使用）"""
    rows = []
    for code, data in history.items():
        name = data.get('name') or STAFF_TABLE.get(code, '')
        for e in data.get('entries', []):
            rows.append({
                '_handler':  code,
                '_name':     name,
                '_desc':     e.get('desc', ''),
                '_amount':   float(e.get('amount', 0)),
                '_date':     e.get('date', ''),
                '_file':     e.get('src_file', ''),
                '_sheet':    e.get('src_sheet', ''),
                '_src_type': e.get('src_type', ''),
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        '_handler','_name','_desc','_amount','_date','_file','_sheet','_src_type'])

# ══════════════════════════════════════════════════
# 3. 查找引擎
# ══════════════════════════════════════════════════

def cjk_substrings(text: str, min_len: int = MIN_MATCH_CHARS):
    cjk = re.sub(r'[^一-鿿㐀-䶿]', '', text)
    for length in range(len(cjk), min_len - 1, -1):
        for start in range(len(cjk) - length + 1):
            yield cjk[start:start + length]

def fuzzy_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()

def keyword_hit(bank_remark: str, ar_desc: str) -> bool:
    """
    從 bank_remark 滑窗取 CJK 子字串（由長到短，最短 MIN_KEYWORD_LEN 字），
    只要有一個出現在 ar_desc 中即命中。
    限制最短 3 字，避免「富邦」「薪資」等泛用詞誤觸。
    """
    if not bank_remark or not ar_desc:
        return False
    for sub in cjk_substrings(bank_remark, min_len=MIN_KEYWORD_LEN):
        if sub in ar_desc:
            return True
    return False

def _search_table(table: pd.DataFrame, bank_remark: str, bank_amount: float,
                  tol: float = None, require_text: bool = False):
    """
    金額篩選 + 關鍵字命中（或 fuzzy fallback）→ 最佳一筆。
    require_text=True：跳過 same-handler shortcut，強制關鍵字/fuzzy 確認（用於歷史資料庫）。
    回傳 row dict 或 None。
    """
    if table.empty or bank_amount <= 0:
        return None
    if tol is None:
        tol = get_tolerance(bank_amount)
    cands = table[abs(table['_amount'] - bank_amount) <= tol].copy()
    if cands.empty:
        return None
    # 精確配對且同一經辦 → 只限 1A（AR 即時表），1B 強制走文字確認
    if tol == 0 and not require_text:
        valid = cands[cands['_handler'].apply(lambda h: bool(str(h).strip()))]
        if not valid.empty and valid['_handler'].nunique() == 1:
            valid = valid.copy()
            valid['_score'] = valid['_desc'].apply(lambda d: fuzzy_score(bank_remark, d))
            top = valid.sort_values('_score', ascending=False).iloc[0]
            if top['_score'] >= EXACT_MIN_FUZZY:
                return top.to_dict()
            # 文字完全不相關，不走 shortcut，繼續進關鍵字/fuzzy 正常路徑
    kw_mask = cands['_desc'].apply(lambda d: keyword_hit(bank_remark, d))
    hits    = cands.loc[kw_mask].copy()
    if hits.empty:
        cands['_score'] = cands['_desc'].apply(lambda d: fuzzy_score(bank_remark, d))
        hits = cands[cands['_score'] >= FUZZY_THRESHOLD].copy()
        if hits.empty:
            return None
        hits = hits.sort_values('_score', ascending=False)
    else:
        hits['_score'] = hits['_desc'].apply(lambda d: fuzzy_score(bank_remark, d))
        hits = hits.sort_values('_score', ascending=False)
    top = hits.iloc[0]
    if not str(top['_handler']).strip():
        return None
    return top.to_dict()

def find_handler(bank_remark: str, bank_amount: float,
                 ar_table: pd.DataFrame, hist_df: pd.DataFrame) -> dict:
    """
    四段式查找：
      精確 1A：未銷帳明細（金額完全吻合）
      緩衝 1A：未銷帳明細（金額±容差 + 文字比對）
      精確 1B：歷史資料庫（金額完全吻合 + 關鍵字/fuzzy 確認）
      緩衝 1B：歷史資料庫（金額±30% + 關鍵字/fuzzy 確認）
      均未找到 → 人工處理
    """
    remark = str(bank_remark or '').strip()
    amount = float(bank_amount) if bank_amount else 0.0
    tol    = get_tolerance(amount)

    def resolve_name(code: str, direct: str = '') -> str:
        if direct:
            return direct
        return STAFF_TABLE.get(code, f'{code}（待確認）') if code else '人工處理'

    def make_result(code, name, method, note,
                    matched_desc='', src_file='', src_sheet='',
                    src_date='', src_amount=None):
        return dict(
            handler_code=code,
            handler_name=resolve_name(code, name),
            method=method,
            note=note,
            matched_desc=matched_desc,
            src_file=src_file,
            src_sheet=src_sheet,
            src_date=src_date,
            src_amount=src_amount,
        )

    # ── 精確 1A：未銷帳明細，金額完全吻合 ────────────
    r = _search_table(ar_table, remark, amount, tol=0)
    if r:
        return make_result(
            r['_handler'], r.get('_name', ''), '1A',
            f'未銷帳明細比對（金額完全吻合，摘要相似度 {round(r.get("_score",0),2)}）',
            r['_desc'], r['_file'], r['_sheet'], r.get('_date', ''), r['_amount'],
        )

    # ── 緩衝 1A：未銷帳明細，金額±容差 ───────────────
    r = _search_table(ar_table, remark, amount, tol=tol)
    if r:
        return make_result(
            r['_handler'], r.get('_name', ''), '1A',
            f'未銷帳明細比對（金額±{tol:,}，摘要相似度 {round(r.get("_score",0),2)}）',
            r['_desc'], r['_file'], r['_sheet'], r.get('_date', ''), r['_amount'],
        )

    # ── 精確 1B：歷史資料庫，金額完全吻合（同一經辦即確定，不需文字）────
    r = _search_table(hist_df, remark, amount, tol=0)
    if r:
        return make_result(
            r['_handler'], r.get('_name', ''), '1B',
            f'歷史資料庫比對（金額完全吻合，摘要相似度 {round(r.get("_score",0),2)}）',
            r['_desc'], r['_file'], r['_sheet'], r.get('_date', ''), r['_amount'],
        )

    # ── 緩衝 1B：歷史資料庫，金額±5% + 文字確認 ────────
    tol_5pct = amount * 0.05
    r = _search_table(hist_df, remark, amount, tol=tol_5pct, require_text=True)
    if r:
        return make_result(
            r['_handler'], r.get('_name', ''), '1B',
            f'歷史資料庫比對（金額±5%，摘要相似度 {round(r.get("_score",0),2)}）',
            r['_desc'], r['_file'], r['_sheet'], r.get('_date', ''), r['_amount'],
        )

    return make_result('', '人工處理', 4, '所有方法均未找到')

# ══════════════════════════════════════════════════
# 4. 更新 Excel：在「待確認」右邊插入「經辦人」欄
# ══════════════════════════════════════════════════

def update_excel(xlsx_path: str, ar_table: pd.DataFrame, hist_df: pd.DataFrame) -> list:
    """
    讀取未入帳清單 Excel，找出標記行，查找經辦人，
    在「待確認」欄右側新增「經辦人」欄（找不到→紅字「人工處理」）。
    回傳所有標記行的結果列表（供 HTML 報告使用）。
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # ── 找表頭行 ──────────────────────────────────
    header_row = None
    col_map    = {}
    for cell_tuple in ws.iter_rows(max_row=15):
        if any('未入帳' in str(c.value or '') for c in cell_tuple):
            header_row = cell_tuple[0].row
            col_map    = {c.column: str(c.value or '') for c in cell_tuple if c.value}
            break

    if header_row is None:
        print(f"  ⚠️  找不到表頭（含「未入帳」），跳過 {xlsx_path}")
        return []

    def find_col(keyword):
        return next((c for c, n in col_map.items() if keyword in n), None)

    remark_col     = find_col('附言')
    amount_in_col  = find_col('存入')
    amount_out_col = find_col('支出')
    date_col       = find_col('日期')
    unmatched_col  = find_col('未入帳')
    pending_col    = find_col('待確認')

    if unmatched_col is None:
        print(f"  ⚠️  找不到「未入帳」欄，跳過")
        return []

    # ── 若「經辦人」欄已存在，先刪除再重建 ────────
    existing = find_col('經辦人')
    if existing:
        ws.delete_cols(existing)
        col_map = {c.column: str(c.value or '')
                   for row in ws.iter_rows(min_row=header_row, max_row=header_row)
                   for c in row if c.value}
        pending_col   = find_col('待確認')
        unmatched_col = find_col('未入帳')

    insert_col = (pending_col + 1) if pending_col else (unmatched_col + 1)
    ws.insert_cols(insert_col)

    # 寫表頭，複製相鄰欄樣式
    ref = ws.cell(row=header_row, column=unmatched_col)
    hdr = ws.cell(row=header_row, column=insert_col, value='經辦人')
    hdr.font      = copy(ref.font)
    hdr.fill      = copy(ref.fill)
    hdr.border    = copy(ref.border)
    hdr.alignment = copy(ref.alignment)
    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width = 12

    # ── 逐行查找 ──────────────────────────────────
    results = []
    for cell_tuple in ws.iter_rows(min_row=header_row + 1):
        r = cell_tuple[0].row
        um_mark = str(ws.cell(row=r, column=unmatched_col).value or '')
        pd_mark = str(ws.cell(row=r, column=pending_col).value or '') if pending_col else ''
        is_unmatched = (um_mark == 'V')
        is_pending   = (pd_mark == 'V')
        if not is_unmatched and not is_pending:
            continue

        remark  = str(ws.cell(row=r, column=remark_col).value or '')  if remark_col     else ''
        amt_in  = parse_amount_str(ws.cell(row=r, column=amount_in_col).value)  if amount_in_col  else 0.0
        amt_out = parse_amount_str(ws.cell(row=r, column=amount_out_col).value) if amount_out_col else 0.0
        amount  = amt_in if amt_in > 0 else amt_out
        date    = str(ws.cell(row=r, column=date_col).value or '') if date_col else ''

        result = find_handler(remark, amount, ar_table, hist_df)

        cell = ws.cell(row=r, column=insert_col)
        ref_data = ws.cell(row=r, column=unmatched_col)
        cell.border    = copy(ref_data.border)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        if result['method'] == 4:
            cell.value = '人工處理'
            cell.font  = Font(bold=True, color='FF0000')
        else:
            cell.value = result['handler_name']
            cell.font  = Font(color='000000')

        results.append({
            'row':          r,
            'date':         date,
            'remark':       remark,
            'amount':       amount,
            'amount_in':    amt_in,
            'amount_out':   amt_out,
            'is_unmatched': is_unmatched,
            'is_pending':   is_pending,
            **result,
        })

    wb.save(xlsx_path)
    print(f"  ✅ 已更新 Excel：{os.path.basename(xlsx_path)}，共 {len(results)} 筆標記")
    return results

# ══════════════════════════════════════════════════
# 5. 產出 HTML 報告
# ══════════════════════════════════════════════════

METHOD_LABEL = {
    '1A': ('🟢', '未銷帳明細'),
    '1B': ('🟢', '歷史資料庫'),
    4:    ('🔴', '人工處理'),
}

def _fmt_src(x: dict) -> str:
    """格式化來源依據欄，供 HTML 使用。"""
    if x['method'] == 4:
        return '<span style="color:#94a3b8">—</span>'
    if x['method'] == 2:
        return f'<span style="color:#1e40af">固定關鍵字規則</span>'

    parts = []
    if x.get('src_file'):
        parts.append(f'<b>{x["src_file"]}</b>')
    if x.get('src_sheet'):
        parts.append(f'工作表：{x["src_sheet"]}')
    if x.get('src_date'):
        parts.append(f'日期：{x["src_date"]}')
    if x.get('src_amount') is not None:
        parts.append(f'金額：{x["src_amount"]:,.0f}')
    if x.get('matched_desc'):
        trunc = x['matched_desc'][:45] + ('…' if len(x['matched_desc']) > 45 else '')
        parts.append(f'摘要：{trunc}')
    return '<br>'.join(parts) if parts else '—'

def build_html_report(all_results: dict, output_path: str):
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    style = """
<style>
  body { font-family:'Noto Sans TC','PingFang TC',system-ui,sans-serif;
         background:#f0f4f8;color:#1e293b;padding:32px; }
  .container { max-width:1500px;margin:auto; }
  h1 { color:#0f172a;border-bottom:4px solid #3b82f6;padding-bottom:12px; }
  h2 { margin-top:40px;color:#1e40af; }
  .summary-grid { display:grid;grid-template-columns:repeat(4,1fr);
                  gap:16px;margin-bottom:32px; }
  .stat-card { background:white;padding:20px;border-radius:12px;
               box-shadow:0 4px 6px rgba(0,0,0,.06);text-align:center; }
  .stat-card .num { font-size:36px;font-weight:800; }
  .month-section { margin-bottom:60px; }
  .month-title { font-size:20px;font-weight:700;padding:10px 16px;
                 background:#1e40af;color:white;border-radius:8px;
                 margin-bottom:20px;display:inline-block; }
  table { width:100%;border-collapse:collapse;font-size:13px;
          background:white;border-radius:12px;overflow:hidden;
          box-shadow:0 4px 6px rgba(0,0,0,.06);margin-bottom:24px; }
  th { background:#f1f5f9;padding:11px 14px;text-align:left;
       border-bottom:2px solid #e2e8f0;font-weight:700;white-space:nowrap; }
  td { padding:10px 14px;border-bottom:1px solid #f1f5f9;vertical-align:top; }
  tr:last-child td { border-bottom:none; }
  tr:hover td { background:#f8fafc; }
  .tag { display:inline-block;padding:3px 10px;border-radius:999px;
         font-size:11px;font-weight:700; }
  .tag-um  { background:#fee2e2;color:#991b1b; }
  .tag-pd  { background:#fef08a;color:#854d0e; }
  .bm1  { background:#dcfce7;color:#166534; }
  .bm2  { background:#dbeafe;color:#1e40af; }
  .bm3  { background:#fef9c3;color:#92400e; }
  .bm4  { background:#fee2e2;color:#991b1b;font-weight:800; }
  .manual { color:#dc2626;font-weight:800; }
  .ain  { color:#16a34a;font-weight:600; }
  .aout { color:#dc2626;font-weight:600; }
  .src-cell { font-size:11.5px;color:#374151;line-height:1.6; }
  .legend { display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap; }
  .legend-item { display:flex;align-items:center;gap:6px;font-size:13px; }
  .footer { margin-top:60px;padding:20px;background:white;border-radius:12px;
            box-shadow:0 4px 6px rgba(0,0,0,.06);font-size:13px;color:#64748b; }
</style>
"""

    total        = sum(len(v) for v in all_results.values())
    manual_count = sum(sum(1 for x in v if x['method'] == 4) for v in all_results.values())
    found_count  = total - manual_count
    m1a = sum(sum(1 for x in v if x['method'] == '1A') for v in all_results.values())
    m1b = sum(sum(1 for x in v if x['method'] == '1B') for v in all_results.values())
    hit_pct = f'{found_count/total*100:.0f}' if total else '0'

    html = f"""<!DOCTYPE html>
<html><head><meta charset='utf-8'>
<title>未入帳經辦人查找報告 ({now})</title>
{style}
</head><body>
<div class='container'>
<h1>📋 未入帳經辦人查找報告</h1>
<p style='color:#64748b'>產出時間：{now}　｜　涵蓋 {len(all_results)} 個月份</p>

<div class='summary-grid'>
  <div class='stat-card' style='border-top:4px solid #3b82f6'>
    <div style='color:#64748b;font-size:13px'>總標記筆數</div>
    <div class='num' style='color:#1e40af'>{total}</div>
  </div>
  <div class='stat-card' style='border-top:4px solid #22c55e'>
    <div style='color:#64748b;font-size:13px'>成功找到經辦</div>
    <div class='num' style='color:#16a34a'>{found_count}</div>
    <div style='font-size:12px;color:#64748b'>{hit_pct}% 命中率</div>
  </div>
  <div class='stat-card' style='border-top:4px solid #ef4444'>
    <div style='color:#64748b;font-size:13px'>需人工處理</div>
    <div class='num' style='color:#dc2626'>{manual_count}</div>
  </div>
  <div class='stat-card' style='border-top:4px solid #64748b'>
    <div style='color:#64748b;font-size:13px'>方法分佈</div>
    <div style='font-size:13px;margin-top:6px;text-align:left'>
      🟢 未銷帳明細比對：{m1a}<br>
      🟢 歷史資料庫比對：{m1b}<br>
      🔴 人工處理：{manual_count}
    </div>
  </div>
</div>

<div class='legend'>
  <div class='legend-item'><span class='tag tag-um'>未入帳</span> 銀行有紀錄、會計尚未入帳</div>
  <div class='legend-item'><span class='tag tag-pd'>待確認</span> 金額相符但語意待確認</div>
  <div class='legend-item'><span class='tag bm1'>🟢 方法1A/1B</span> 金額容差+關鍵字/fuzzy 比對</div>
  <div class='legend-item'><span class='tag bm4'>🔴 人工處理</span> 兩種方法均未找到</div>
</div>
"""

    for month_code, items in sorted(all_results.items()):
        if not items:
            continue
        m_found  = sum(1 for x in items if x['method'] != 4)
        m_manual = sum(1 for x in items if x['method'] == 4)
        html += f"""
<div class='month-section'>
<div class='month-title'>📅 {month_code}　共 {len(items)} 筆　｜ 找到 {m_found} · 人工 {m_manual}</div>
<table>
<thead><tr>
  <th>#</th>
  <th>銀行日期</th>
  <th>銀行附言</th>
  <th>存入</th>
  <th>支出</th>
  <th>類型</th>
  <th>查找方法</th>
  <th>經辦人</th>
  <th>來源依據</th>
</tr></thead>
<tbody>
"""
        badge_map = {'1A':'bm1','1B':'bm1',4:'bm4'}
        for i, x in enumerate(items, 1):
            badge_cls = badge_map.get(x['method'], '')
            emoji, label = METHOD_LABEL.get(x['method'], ('', ''))
            type_tag = ("<span class='tag tag-um'>未入帳</span>" if x['is_unmatched']
                        else "<span class='tag tag-pd'>待確認</span>")
            name_cell = (f"<span class='manual'>⚠ {x['handler_name']}</span>"
                         if x['method'] == 4
                         else f"<b>{x['handler_name']}</b>"
                              + (f"<br><span style='color:#94a3b8;font-size:11px'>{x['handler_code']}</span>"
                                 if x['handler_code'] else ''))
            ain_str  = f"<span class='ain'>+{x['amount_in']:,.0f}</span>"  if x['amount_in']  > 0 else '—'
            aout_str = f"<span class='aout'>-{x['amount_out']:,.0f}</span>" if x['amount_out'] > 0 else '—'
            src_html = _fmt_src(x)
            html += f"""<tr>
  <td style='color:#94a3b8'>{i}</td>
  <td style='white-space:nowrap'>{x['date']}</td>
  <td>{x['remark']}</td>
  <td style='text-align:right'>{ain_str}</td>
  <td style='text-align:right'>{aout_str}</td>
  <td>{type_tag}</td>
  <td><span class='tag {badge_cls}'>{emoji} {label}</span><br>
      <span style='font-size:10px;color:#64748b'>{x['note']}</span></td>
  <td>{name_cell}</td>
  <td class='src-cell'>{src_html}</td>
</tr>"""
        html += "</tbody></table></div>"

    html += f"""
<div class='footer'>
  <b>查找方法說明</b><br>
  🟢 <b>方法 1A（未銷帳明細）</b>：搜尋即時的未銷帳明細表_.xlsm。金額容差（≤10,000 用±500；其餘±2,000）+ 關鍵字/fuzzy 比對。<br>
  🟢 <b>方法 1B（歷史資料庫）</b>：搜尋 handler_history.json（累積的帳務查詢 + 未銷帳明細）。同金額容差 + 關鍵字/fuzzy 比對。<br>
  兩者均先嘗試關鍵字子字串命中，再以 fuzzy 相似度排序取最高分。<br>
  🔴 <b>人工處理</b>：兩種方法均未找到，需打銀行或內部確認。<br><br>
  歷史資料庫路徑：{HISTORY_PATH}
</div>
</div></body></html>
"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  ✅ HTML 報告：{output_path}")

# ══════════════════════════════════════════════════
# 6. 主流程
# ══════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("步驟 1：更新歷史資料庫")
    print("=" * 55)
    history = update_handler_history()

    print("\n" + "=" * 55)
    print("步驟 2：載入即時查找表")
    print("=" * 55)
    print("載入未銷帳明細表（即時）…")
    ar_table = load_ar_table()
    print(f"  共 {len(ar_table)} 筆（{ar_table['_sheet'].nunique()} 個科目）")

    print("展開歷史資料庫…")
    hist_df = history_to_df(history)
    print(f"  共 {len(hist_df)} 筆歷史記錄（{hist_df['_handler'].nunique()} 位經辦）")

    print("\n" + "=" * 55)
    print("步驟 3：處理未入帳清單")
    print("=" * 55)
    xlsx_files = sorted(glob.glob(os.path.join(DATA_DIR, '未入帳清單_*.xlsx')))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith('~$')]

    if not xlsx_files:
        print(f"在 {DATA_DIR} 找不到未入帳清單 xlsx 檔案")
        return

    print(f"找到 {len(xlsx_files)} 個月份的未入帳清單")

    all_results = {}
    for path in xlsx_files:
        month = os.path.basename(path).replace('未入帳清單_', '').replace('.xlsx', '')
        print(f"\n處理 {month}…")
        items = update_excel(path, ar_table, hist_df)
        all_results[month] = items

    print("\n" + "=" * 55)
    print("步驟 4：產出 HTML 報告")
    print("=" * 55)
    html_path = os.path.join(OUTPUT_DIR, '未入帳_經辦人報告.html')
    build_html_report(all_results, html_path)

    total = sum(len(v) for v in all_results.values())
    found = sum(sum(1 for x in v if x['method'] != 4) for v in all_results.values())
    print(f"\n{'='*55}")
    print(f"完成！總計 {total} 筆，找到經辦 {found} 筆，人工處理 {total-found} 筆")
    print(f"HTML 報告：{html_path}")

if __name__ == '__main__':
    main()
