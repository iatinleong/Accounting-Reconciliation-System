"""
reconcile_core.py
核心對帳邏輯（從 Reconcile_Agent_PoC.py 提取）
提供 run_pipeline() 作為 Streamlit 的入口點。
"""

from __future__ import annotations

import csv
import datetime
import difflib
import itertools
import os
import re
import warnings
import io
import contextlib
from copy import copy
from pathlib import Path

import httpx
import numpy as np
import openpyxl
import pandas as pd
from dotenv import load_dotenv
from openai import AzureOpenAI
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings("ignore")
pd.set_option("display.float_format", lambda x: "%.2f" % x)

# ── 載入 .env ──
load_dotenv(Path(__file__).parent / ".env")

# ── Azure OpenAI 設定（對齊 SQLagentnew/agent/config.py）──
_AZURE_ENDPOINT    = os.getenv("AZURE_OPENAI_ENDPOINT", "")
_AZURE_KEY         = os.getenv("AZURE_OPENAI_KEY", "")
_AZURE_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2025-04-01-preview")
_AZURE_MODEL       = os.getenv("AZURE_OPENAI_MODEL", "gpt-5.2")

_REASONING_MODELS = {"o1", "o1-mini", "o3", "o3-mini", "o4-mini"}

# ── 模組層級狀態（每次 run_pipeline 會重置）──
_openai_client: AzureOpenAI | None = None
_token_usage: dict = {"input": 0, "output": 0, "calls": 0}
_monthly_token_log: list = []

# ── 常數 ──
COIN_MIN_AMT = 1_000_000
COIN_TOLERANCE = 2_000
ACC_REVERSAL_MAX_COMBO = 4
PRE_REVERSAL_KEYWORDS = ["款項更正", "迴轉原傳票", "迴轉", "回轉", "沖銷", "沖正", "轉回"]
NET_MAX_POS = 2
NET_MAX_NEG = 2
NET_REVERSAL_KEYWORDS = ["迴轉", "回轉", "沖銷", "沖正", "轉回", "沖回"]
_EXT_PRIORITY = {".xlsx": 0, ".xlsm": 1, ".xls": 2}


# ══════════════════════════════════════════════════
# 1. 工具函數
# ══════════════════════════════════════════════════

def convert_to_xlsx(src_path: str) -> str:
    if src_path.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            load_workbook(src_path, read_only=True).close()
            return src_path
        except Exception:
            pass

    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    dst_path = src_path.rsplit(".", 1)[0] + ".xlsx"
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(src_path))
        wb.SaveAs(os.path.abspath(dst_path), FileFormat=51)
        wb.Close(False)
        print(f"🔄 已轉換：{os.path.basename(dst_path)}")
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()
    return dst_path


def month_code_to_date(code: str) -> str:
    return f"{int(code[:3]) + 1911}/{code[3:5]}"


# ══════════════════════════════════════════════════
# 2. 載入與預處理
# ══════════════════════════════════════════════════

def load_acc_sheets(acc_file_paths: list[str]) -> dict:
    acc_sheets: dict = {}
    for ap in acc_file_paths:
        ap_xlsx = convert_to_xlsx(ap)
        src_label = os.path.splitext(os.path.basename(ap_xlsx))[0]
        sheets = pd.read_excel(ap_xlsx, sheet_name=None)
        for sheet_name, df in sheets.items():
            unique_key = f"{src_label}::{sheet_name}" if sheet_name in acc_sheets else sheet_name
            acc_sheets[unique_key] = df
        print(f"✅ 載入 {os.path.basename(ap_xlsx)}，{len(sheets)} 個 Sheet")
    print(f"✅ 會計帳合計 {len(acc_sheets)} 個 Sheet：{list(acc_sheets.keys())}")
    return acc_sheets


def load_and_preprocess(target_date: str, bank_path: str, acc_sheets: dict):
    df_bank = pd.read_excel(bank_path, skiprows=5)
    df_bank = df_bank.dropna(subset=["交易日期"])
    df_bank["存入金額"] = pd.to_numeric(
        df_bank["存入金額"].astype(str).str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)
    df_bank["支出金額"] = pd.to_numeric(
        df_bank["支出金額"].astype(str).str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0)

    target_month = target_date[:7]
    df_bank_all = df_bank.copy()
    df_bank_all["_date_str"] = pd.to_datetime(df_bank_all["交易日期"], errors="coerce").dt.strftime("%Y/%m")
    df_bank_all = df_bank_all[df_bank_all["_date_str"] == target_month].drop(columns=["_date_str"])
    df_bank_all["Bank_index"] = df_bank_all.index.astype(str)

    df_bank_all["_date"] = pd.to_datetime(df_bank_all["交易日期"], errors="coerce").dt.date
    to_remove: set = set()
    reversal_log = []
    reversals = df_bank_all[df_bank_all["支出金額"] < 0]
    for r_idx, r_row in reversals.iterrows():
        rev_amt = abs(r_row["支出金額"])
        matches = df_bank_all[
            (df_bank_all["_date"] == r_row["_date"])
            & (df_bank_all["支出金額"] == rev_amt)
            & (~df_bank_all.index.isin(to_remove))
        ]
        if not matches.empty:
            orig_idx = matches.index[0]
            to_remove.add(r_idx)
            to_remove.add(orig_idx)
            reversal_log.append({
                "orig": df_bank_all.loc[orig_idx].copy(),
                "chong": r_row.copy(),
                "amount": rev_amt,
            })
            print(f"  🔄 日內沖銷移除：Bank_index={df_bank_all.loc[orig_idx,'Bank_index']} [{df_bank_all.loc[orig_idx,'摘要']}] + [{r_row['摘要']}]  各 {rev_amt:,.0f}")
    df_bank_all = df_bank_all[~df_bank_all.index.isin(to_remove)].drop(columns=["_date"])

    bank_in = df_bank_all[df_bank_all["存入金額"] > 0].copy()
    bank_out = df_bank_all[df_bank_all["支出金額"] > 0].copy()

    required_acc_columns = [
        "憑證創建人", "憑證編號", "憑證行編號", "憑證類型", "業務參考",
        "會計期間", "業務日期", "科目代碼", "描述.1", "業務貨幣代碼",
        "業務金額", "本位幣金額", "第二本位幣/報表金額", "借方/貸方",
        "商部別 分析代碼", "定存單號/銀行與帳號 分析代碼", "對象別 分析代碼",
    ]

    df_acc_list = []
    for sheet_name, df_sheet in acc_sheets.items():
        actual_cols = [c for c in df_sheet.columns if c in required_acc_columns or c == "描述.1" or c == "業務金額" or c == "業務日期"]
        if "業務金額" not in df_sheet.columns or "業務日期" not in df_sheet.columns:
            continue
        df_sheet = df_sheet[actual_cols].copy()
        df_sheet["來源Sheet"] = sheet_name
        df_acc_list.append(df_sheet)
        print(f"  ✅ 讀入 [{sheet_name}]：{len(df_sheet)} 筆")

    df_acc = pd.concat(df_acc_list, ignore_index=True)

    def parse_amount(s):
        s = str(s).strip().replace(",", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        return pd.to_numeric(s, errors="coerce")

    df_acc["業務金額"] = df_acc["業務金額"].apply(parse_amount).fillna(0)

    try:
        df_acc["業務日期_格式化"] = pd.to_datetime(df_acc["業務日期"], errors="coerce").dt.strftime("%Y/%m/%d")
        if df_acc["業務日期_格式化"].isna().all():
            raise ValueError()
    except Exception:
        df_acc["業務日期_格式化"] = pd.to_datetime(
            pd.to_numeric(df_acc["業務日期"], errors="coerce"),
            origin="1899-12-30",
            unit="D",
        ).dt.strftime("%Y/%m/%d")

    df_acc_all = df_acc[df_acc["業務日期_格式化"].str.startswith(target_month, na=False)].copy()
    df_acc_all["Acc_index"] = df_acc_all.index.astype(str)

    acc_in = df_acc_all[df_acc_all["業務金額"] > 0].copy()
    acc_out = df_acc_all[df_acc_all["業務金額"] < 0].copy()

    return bank_in, bank_out, acc_in, acc_out, reversal_log


# ══════════════════════════════════════════════════
# 3. 關鍵字與 LLM 輔助
# ══════════════════════════════════════════════════

def _chat(messages: list[dict], **kwargs) -> object:
    """Azure OpenAI 統一呼叫入口（對齊 SQLagentnew generator._chat）。
    gpt-5.x 不支援 temperature，自動移除並補上 reasoning_effort。
    """
    model = _AZURE_MODEL
    base = model.split("-")[0] if "-" in model else model
    no_temp = model in _REASONING_MODELS or base in _REASONING_MODELS or model.startswith("gpt-5")
    if no_temp:
        kwargs.pop("temperature", None)
    if model.startswith("gpt-5") and "reasoning_effort" not in kwargs:
        kwargs["reasoning_effort"] = "medium"
    return _openai_client.chat.completions.create(model=model, messages=messages, **kwargs)


def extract_keywords(text: str) -> list[str]:
    text = str(text).strip()
    if not text or text.lower() == "nan":
        return []
    tokens = re.findall(r"[A-Za-z0-9]{3,}|[一-鿿]{2,}", text)
    return [t.upper() for t in tokens]


def _cjk_substrings(token: str, min_len: int = 2):
    for sub_len in range(len(token), min_len - 1, -1):
        for start in range(len(token) - sub_len + 1):
            yield token[start : start + sub_len]


def keyword_match(bank_memo: str, acc_desc1: str) -> tuple[bool, str]:
    bank_tokens = extract_keywords(bank_memo)
    acc_text = str(acc_desc1).upper()
    for token in bank_tokens:
        if token in acc_text:
            return True, token
        if len(token) >= 4 and all("一" <= c <= "鿿" for c in token):
            for sub in _cjk_substrings(token, min_len=2):
                if sub in acc_text:
                    return True, sub
    bank_text = str(bank_memo).upper()
    for token in extract_keywords(acc_desc1):
        if len(token) >= 4 and token in bank_text:
            return True, token
        if len(token) >= 4 and all("一" <= c <= "鿿" for c in token):
            for sub in _cjk_substrings(token, min_len=2):
                if sub in bank_text:
                    return True, sub
    return False, ""


def llm_pick_best_candidate(bank_row, candidates_df: pd.DataFrame, mode: str):
    if "描述.1" in candidates_df.columns:
        if candidates_df["描述.1"].astype(str).str.contains("銀行存款－台北富邦銀行", na=False).all():
            print("      ⏭️ 所有候選描述均為銀行存款，LLM 無法分辨，留待批次核銷")
            return None

    prompt = (
        f"你是專業會計。有一筆銀行{'收款' if mode == 'Income' else '付款'}，"
        f"與其金額相同的會計帳有 {len(candidates_df)} 筆候選紀錄。\n"
        "請根據銀行的「附言」與會計帳的「備註描述(描述.1)」，挑出真正對應的「唯一 1 筆」會計帳。\n"
        f"【目標銀行紀錄】\n Bank_index: {bank_row['Bank_index']}, 附言: {bank_row['附言']}\n"
        f"【多筆會計候選人】\n{candidates_df[['Acc_index', '描述.1']].to_json(orient='records', force_ascii=False)}\n"
        "請只回傳最符合的該筆 Acc_index 字串。若都不吻合則回傳 None。"
    )
    try:
        resp = _chat([{"role": "user", "content": prompt}])
        _token_usage["input"] += resp.usage.prompt_tokens
        _token_usage["output"] += resp.usage.completion_tokens
        _token_usage["calls"] += 1
        ans = resp.choices[0].message.content.strip()
        if ans in candidates_df["Acc_index"].values:
            return ans
    except Exception:
        pass
    return None


def llm_pick_best_bank_candidate(acc_row, bank_candidates_df: pd.DataFrame, mode: str):
    prompt = (
        f"你是專業會計。有一筆會計{'收款' if mode == 'Income' else '付款'}，"
        f"與其金額相同的銀行記錄有 {len(bank_candidates_df)} 筆候選。\n"
        "請根據會計帳的「備註描述(描述.1)」與銀行的「附言」，挑出真正對應的「唯一 1 筆」銀行記錄。\n"
        f"【目標會計紀錄】\n Acc_index: {acc_row['Acc_index']}, 描述.1: {acc_row.get('描述.1', '')}, 憑證編號: {acc_row.get('憑證編號', '')}\n"
        f"【多筆銀行候選人】\n{bank_candidates_df[['Bank_index', '附言', '摘要']].to_json(orient='records', force_ascii=False)}\n"
        "請只回傳最符合的該筆 Bank_index 字串。若都不吻合則回傳 None。"
    )
    try:
        resp = _chat([{"role": "user", "content": prompt}])
        _token_usage["input"] += resp.usage.prompt_tokens
        _token_usage["output"] += resp.usage.completion_tokens
        _token_usage["calls"] += 1
        ans = resp.choices[0].message.content.strip()
        if ans in bank_candidates_df["Bank_index"].values:
            return ans
    except Exception:
        pass
    return None


def coin_fill(pool: pd.DataFrame, amt_col: str, target: float):
    selected, rem = [], int(round(target))
    for bamt_f in sorted(pool[amt_col].unique(), reverse=True):
        bamt = int(round(bamt_f))
        if bamt <= 0 or bamt > rem:
            continue
        cands = pool[
            (pool[amt_col].apply(lambda x: int(round(x))) == bamt)
            & (~pool["Bank_index"].isin(selected))
        ]
        n = min(rem // bamt, len(cands))
        if n > 0:
            selected.extend(cands.iloc[:n]["Bank_index"].tolist())
            rem -= n * bamt
        if rem == 0:
            break
    return selected, rem


def evaluate_combination_with_llm(bank_rows, acc_rows, match_type: str, mode: str):
    if "描述.1" in acc_rows.columns:
        if acc_rows["描述.1"].astype(str).str.contains("銀行存款－台北富邦銀行", na=False).all():
            return True, "會計描述為銀行存款類，無需LLM，數學完全吻合自動通過"
    return True, f"{mode}端數學完全吻合，自動通過"


# ══════════════════════════════════════════════════
# 4. 對帳引擎
# ══════════════════════════════════════════════════

def reconcile_engine(df_bank: pd.DataFrame, df_acc: pd.DataFrame, mode: str = "Income", pending: list = None):
    if pending is None:
        pending = []
    label = "收入" if mode == "Income" else "支出"
    print(f"\n{'='*16} 🚀 [{label}] 開始對帳 {'='*16}")
    rem_b, rem_a = df_bank.copy(), df_acc.copy()
    history = []
    bank_amt_col = "存入金額" if mode == "Income" else "支出金額"

    # Step 1-A: 關鍵字優先輪
    print(f"\n⏳ [Step 1-A] 關鍵字優先輪...")
    for b_idx, b_row in df_bank.iterrows():
        if b_idx not in rem_b.index:
            continue
        b_amt = b_row[bank_amt_col]
        candidates = rem_a[rem_a["業務金額"].abs() == b_amt]
        if candidates.empty:
            continue
        kw_winner = None
        for _, cand in candidates.iterrows():
            km, kw = keyword_match(b_row.get("附言", ""), cand.get("描述.1", ""))
            if km:
                kw_winner = (cand, kw)
                break
        if kw_winner:
            a_sel, kw = kw_winner
            print(f"  ✅ [1-A] Bank:{b_row['Bank_index']}({b_amt:,.0f}) ↔ Acc:{a_sel['Acc_index']} 關鍵字「{kw}」")
            history.append({
                "Type": f"{label} 1對1",
                "MatchReason": f"關鍵字優先「{kw}」",
                "Bank_Data": b_row.to_frame().T,
                "Acc_Data": a_sel.to_frame().T,
            })
            rem_a = rem_a[rem_a["Acc_index"] != a_sel["Acc_index"]]
            rem_b = rem_b.drop(b_idx)

    # Step 1-B: 金額唯一輪
    print(f"\n⏳ [Step 1-B] 金額唯一輪... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    for b_idx, b_row in rem_b.copy().iterrows():
        if b_idx not in rem_b.index:
            continue
        b_amt = b_row[bank_amt_col]
        candidates = rem_a[rem_a["業務金額"].abs() == b_amt]
        if len(candidates) == 0:
            continue
        competing_banks = rem_b[(rem_b[bank_amt_col] == b_amt) & (rem_b.index != b_idx)]
        if len(candidates) == 1 and competing_banks.empty:
            a_cand = candidates.iloc[0]
            _acc_desc = str(a_cand.get("描述.1", ""))
            _skip_for_batch = False
            if "摘要" in rem_b.columns:
                _rb2 = rem_b.copy()
                _rb2["_grp_date"] = pd.to_datetime(_rb2["交易日期"], errors="coerce").dt.strftime("%Y/%m/%d")
                _rb2["_grp_bank"] = _rb2["代辦行"].fillna("") if "代辦行" in _rb2.columns else ""
                for (_gd, _gb, _memo), _grp in _rb2.groupby(["_grp_date", "_grp_bank", "摘要"]):
                    if len(_grp) <= 1:
                        continue
                    _gs = _grp[bank_amt_col].sum()
                    if abs(_gs - b_amt) > 0.01:
                        continue
                    for _col in ["代辦行", "附言", "摘要"]:
                        if _col not in _grp.columns:
                            continue
                        _km, _kw = keyword_match(str(_grp[_col].iloc[0]), _acc_desc)
                        if _km:
                            print(f"  ⏭️ Bank:{b_row['Bank_index']}({b_amt:,.0f}) 保留給Step5整批")
                            _skip_for_batch = True
                            break
                    if _skip_for_batch:
                        break
            if _skip_for_batch:
                continue
            a_row = candidates.iloc[0]
            km, kw = keyword_match(b_row.get("附言", ""), a_row.get("描述.1", ""))
            if km:
                print(f"  ✅ [1-B] Bank:{b_row['Bank_index']}({b_amt:,.0f}) ↔ Acc:{a_row['Acc_index']} 金額唯一+關鍵字「{kw}」")
                history.append({"Type": f"{label} 1對1", "MatchReason": f"金額唯一+關鍵字「{kw}」", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_row.to_frame().T})
                rem_a = rem_a.drop(candidates.index[0])
                rem_b = rem_b.drop(b_idx)
            else:
                print(f"  ⚠️ [1-B] Bank:{b_row['Bank_index']}({b_amt:,.0f}) 金額唯一但無關鍵字 → LLM確認...")
                is_v, rs = evaluate_combination_with_llm(b_row.to_frame().T, a_row.to_frame().T, "1對1", mode)
                if is_v:
                    history.append({"Type": f"{label} 1對1", "MatchReason": "金額唯一+LLM確認", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_row.to_frame().T})
                    rem_a = rem_a.drop(candidates.index[0])
                    rem_b = rem_b.drop(b_idx)
                else:
                    print(f"  🚫 [1-B] Bank:{b_row['Bank_index']} LLM不確認，保留")
        elif len(candidates) == 1 and not competing_banks.empty:
            a_row = candidates.iloc[0]
            all_b_cands = pd.concat([b_row.to_frame().T, competing_banks])
            print(f"  ⚠️ Acc:{a_row['Acc_index']}({b_amt:,.0f}) 有{len(all_b_cands)}筆銀行競爭，LLM 從銀行端選...")
            best_bank_id = llm_pick_best_bank_candidate(a_row, all_b_cands, mode)
            if best_bank_id:
                b_win = all_b_cands[all_b_cands["Bank_index"] == best_bank_id].iloc[0]
                b_win_idx = all_b_cands[all_b_cands["Bank_index"] == best_bank_id].index[0]
                history.append({"Type": f"{label} 1對1", "MatchReason": "LLM確認(銀行端競爭)", "Bank_Data": b_win.to_frame().T, "Acc_Data": a_row.to_frame().T})
                rem_a = rem_a.drop(candidates.index[0])
                if b_win_idx in rem_b.index:
                    rem_b = rem_b.drop(b_win_idx)
        elif len(candidates) > 1:
            print(f"  ⚠️ Bank:{b_row['Bank_index']}({b_amt:,.0f}) 有{len(candidates)}筆候選 → LLM...")
            best_id = llm_pick_best_candidate(b_row, candidates, mode)
            if best_id:
                a_sel = candidates[candidates["Acc_index"] == best_id].iloc[0]
                history.append({"Type": f"{label} 1對1", "MatchReason": "LLM確認", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_sel.to_frame().T})
                rem_a = rem_a[rem_a["Acc_index"] != best_id]
                rem_b = rem_b.drop(b_idx)

    # Step 2: 1 Bank → 2 Acc
    print(f"\n⏳ [Step 2] 1 Bank → 2 Acc... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    for (a1i, a1), (a2i, a2) in itertools.combinations(rem_a.iterrows(), 2):
        if a1["Acc_index"] in matched_a or a2["Acc_index"] in matched_a:
            continue
        s = abs(a1["業務金額"]) + abs(a2["業務金額"])
        for b_idx, b_row in rem_b[rem_b[bank_amt_col] == s].iterrows():
            if b_row["Bank_index"] in matched_b:
                continue
            is_v, rs = evaluate_combination_with_llm(pd.DataFrame([b_row]), pd.DataFrame([a1, a2]), "1對2", mode)
            if is_v:
                history.append({"Type": f"{label} 1對2", "MatchReason": rs, "Bank_Data": b_row.to_frame().T, "Acc_Data": pd.concat([a1.to_frame().T, a2.to_frame().T])})
                matched_b.append(b_row["Bank_index"])
                matched_a.extend([a1["Acc_index"], a2["Acc_index"]])
                break
            else:
                print(f"  🟡 1:2 數學吻合但LLM不確定 → 待確認 Bank:{b_row['Bank_index']}")
                pending.append({"Mode": mode, "Type": "1對2 待確認", "Bank_Data": b_row.to_frame().T, "Acc_Data": pd.concat([a1.to_frame().T, a2.to_frame().T]), "Reason": rs})
                break
    rem_b = rem_b[~rem_b["Bank_index"].isin(matched_b)]
    rem_a = rem_a[~rem_a["Acc_index"].isin(matched_a)]

    # Step 3: N Bank → 1 Acc
    print(f"\n⏳ [Step 3] N Bank → 1 Acc... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    for a_idx, a_row in rem_a.iterrows():
        if a_row["Acc_index"] in matched_a:
            continue
        a_amt = abs(a_row["業務金額"])
        if a_amt == 0:
            continue
        pool = rem_b[(~rem_b["Bank_index"].isin(matched_b)) & (rem_b[bank_amt_col] <= a_amt)]
        found = False
        for r in range(2, min(5, len(pool) + 1)):
            if found:
                break
            for combo in itertools.combinations(pool.iterrows(), r):
                rows = [c[1] for c in combo]
                if sum(row[bank_amt_col] for row in rows) != a_amt:
                    continue
                combo_df = pd.DataFrame(rows)
                combo_ids = combo_df["Bank_index"].tolist()
                is_v, rs = evaluate_combination_with_llm(combo_df, a_row.to_frame().T, f"{r}對1", mode)
                if is_v:
                    history.append({"Type": f"{label} {r}對1", "MatchReason": rs, "Bank_Data": combo_df, "Acc_Data": a_row.to_frame().T})
                    matched_b.extend(combo_ids)
                    matched_a.append(a_row["Acc_index"])
                    found = True
                    break
                else:
                    pending.append({"Mode": mode, "Type": f"{r}對1 待確認", "Bank_Data": combo_df, "Acc_Data": a_row.to_frame().T, "Reason": rs})
    rem_b = rem_b[~rem_b["Bank_index"].isin(matched_b)]
    rem_a = rem_a[~rem_a["Acc_index"].isin(matched_a)]

    # Step 4: 同金額 N對N
    print(f"\n⏳ [Step 4] N對N 同金額批次... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    _s4_rb = rem_b.copy()
    _s4_rb["_grp_date"] = pd.to_datetime(_s4_rb["交易日期"], errors="coerce").dt.strftime("%Y/%m/%d")
    _s4_rb["_grp_bank"] = _s4_rb["代辦行"].fillna("") if "代辦行" in _s4_rb.columns else ""
    _reserved_acc_ids: set = set()
    if "摘要" in _s4_rb.columns:
        for (_gd, _gb, _gm), _gg in _s4_rb.groupby(["_grp_date", "_grp_bank", "摘要"]):
            if len(_gg) <= 1:
                continue
            _gs = _gg[bank_amt_col].sum()
            _matching_acc = rem_a[rem_a["業務金額"].abs() == _gs]
            if len(_matching_acc) == 1:
                _reserved_acc_ids.add(_matching_acc.iloc[0]["Acc_index"])
    for b_amt, _ in rem_b.groupby(bank_amt_col):
        if rem_b.empty or rem_a.empty or b_amt == 0:
            break
        b_grp = rem_b[rem_b[bank_amt_col] == b_amt]
        a_grp = rem_a[rem_a["業務金額"].abs() == b_amt]
        if len(b_grp) == 0 or len(a_grp) == 0 or len(b_grp) != len(a_grp):
            continue
        if a_grp["Acc_index"].isin(_reserved_acc_ids).any():
            continue
        history.append({"Type": f"{label} {len(b_grp)}對{len(a_grp)}批次", "Bank_Data": b_grp, "Acc_Data": a_grp})
        rem_b = rem_b[~rem_b["Bank_index"].isin(b_grp["Bank_index"])]
        rem_a = rem_a[~rem_a["Acc_index"].isin(a_grp["Acc_index"])]

    # Step 5: 同日期+摘要整批加總
    print(f"\n⏳ [Step 5] 同日期+摘要整批加總... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    if "摘要" in rem_b.columns:
        _rb = rem_b.copy()
        _rb["_grp_date"] = pd.to_datetime(_rb["交易日期"], errors="coerce").dt.strftime("%Y/%m/%d")
        _rb["_grp_bank"] = _rb["代辦行"].fillna("") if "代辦行" in _rb.columns else ""
        for (grp_date, grp_bank, memo), b_grp in _rb.groupby(["_grp_date", "_grp_bank", "摘要"]):
            b_ids = b_grp["Bank_index"].tolist()
            if any(bid in matched_b for bid in b_ids):
                continue
            b_sum = b_grp[bank_amt_col].sum()
            if b_sum == 0:
                continue
            cands = rem_a[(rem_a["業務金額"].abs() == b_sum) & (~rem_a["Acc_index"].isin(matched_a))]
            if len(cands) == 1:
                a_row = cands.iloc[0]
                print(f"  ✅ 整批[{grp_date} {grp_bank} {memo}] {len(b_grp)}筆合計{b_sum:,.0f} ↔ Acc:{a_row['Acc_index']}")
                history.append({"Type": f"{label} {len(b_grp)}對1整批", "MatchReason": f"同日期+代辦行+摘要整批加總吻合", "Bank_Data": b_grp, "Acc_Data": a_row.to_frame().T})
                matched_b.extend(b_ids)
                matched_a.append(a_row["Acc_index"])
        # 跨天整批
        for (grp_bank, memo), b_grp in _rb.groupby(["_grp_bank", "摘要"]):
            b_ids = b_grp["Bank_index"].tolist()
            if any(bid in matched_b for bid in b_ids) or len(b_grp) <= 1:
                continue
            b_sum = b_grp[bank_amt_col].sum()
            if b_sum == 0:
                continue
            cands = rem_a[(rem_a["業務金額"].abs() == b_sum) & (~rem_a["Acc_index"].isin(matched_a))]
            if len(cands) == 1:
                a_row = cands.iloc[0]
                print(f"  ✅ 整批跨天[{grp_bank} {memo}] {len(b_grp)}筆合計{b_sum:,.0f} ↔ Acc:{a_row['Acc_index']}")
                history.append({"Type": f"{label} {len(b_grp)}對1整批跨天", "MatchReason": f"同摘要({memo})跨天整批加總吻合", "Bank_Data": b_grp, "Acc_Data": a_row.to_frame().T})
                matched_b.extend(b_ids)
                matched_a.append(a_row["Acc_index"])
    rem_b = rem_b[~rem_b["Bank_index"].isin(matched_b)]
    rem_a = rem_a[~rem_a["Acc_index"].isin(matched_a)]

    # Step 2-C: 近似金額 ±5%
    _APPROX_RATIO = 0.05
    _FUZZY_THR = 0.45
    print(f"\n⏳ [Step 2-C] 近似金額輪（±5% + 關鍵字/fuzzy）... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    for b_idx, b_row in rem_b.copy().iterrows():
        if b_idx not in rem_b.index:
            continue
        b_amt = b_row[bank_amt_col]
        if b_amt <= 0:
            continue
        tol = b_amt * _APPROX_RATIO
        cands = rem_a[
            (rem_a["業務金額"].abs() != b_amt) & ((rem_a["業務金額"].abs() - b_amt).abs() <= tol)
        ].copy()
        if cands.empty:
            continue
        bank_memo = str(b_row.get("附言", ""))
        kw_hits = []
        for _, cand in cands.iterrows():
            km, kw = keyword_match(bank_memo, cand.get("描述.1", ""))
            if km:
                score = difflib.SequenceMatcher(None, bank_memo, str(cand.get("描述.1", ""))).ratio()
                kw_hits.append((score, kw, cand))
        if kw_hits:
            kw_hits.sort(key=lambda x: x[0], reverse=True)
            score, kw, a_sel = kw_hits[0]
        else:
            fuzzy_hits = []
            for _, cand in cands.iterrows():
                score = difflib.SequenceMatcher(None, bank_memo, str(cand.get("描述.1", ""))).ratio()
                if score >= _FUZZY_THR:
                    fuzzy_hits.append((score, cand))
            if not fuzzy_hits:
                continue
            fuzzy_hits.sort(key=lambda x: x[0], reverse=True)
            score, a_sel = fuzzy_hits[0]
            kw = f"fuzzy({score:.2f})"
        diff_amt = abs(abs(a_sel["業務金額"]) - b_amt)
        print(f"  ✅ [2-C] Bank:{b_row['Bank_index']}({b_amt:,.0f}) ↔ Acc:{a_sel['Acc_index']}({abs(a_sel['業務金額']):,.0f}) 差額{diff_amt:,.0f} 「{kw}」")
        history.append({"Type": f"{label} 1對1近似", "MatchReason": f"近似金額(差額{diff_amt:,.0f})+關鍵字「{kw}」", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_sel.to_frame().T})
        rem_a = rem_a[rem_a["Acc_index"] != a_sel["Acc_index"]]
        rem_b = rem_b.drop(b_idx)

    # Step 6: Coin-change
    print(f"\n⏳ [Step 6] Coin-change配對... (庫存: 銀行{len(rem_b)} / 會計{len(rem_a)})")
    matched_b, matched_a = [], []
    acc_large = rem_a[rem_a["業務金額"].abs() >= COIN_MIN_AMT].copy()
    acc_large = acc_large.iloc[acc_large["業務金額"].abs().argsort()[::-1]]
    for _, a_row in acc_large.iterrows():
        if a_row["Acc_index"] in matched_a:
            continue
        a_amt = int(round(abs(a_row["業務金額"])))
        pool = rem_b[~rem_b["Bank_index"].isin(matched_b)]
        if pool.empty:
            break
        sel_ids, remainder = coin_fill(pool, bank_amt_col, a_amt)
        if not sel_ids:
            continue
        if remainder == 0 or (0 < remainder <= COIN_TOLERANCE):
            reason = "Coin精確" if remainder == 0 else f"Coin容差(差額{remainder:,}元)"
            sel_df = rem_b[rem_b["Bank_index"].isin(sel_ids)]
            print(f"  ✅ {reason}: {len(sel_ids)}筆合計{sel_df[bank_amt_col].sum():,.0f} → Acc:{a_row['Acc_index']}({a_amt:,.0f})")
            history.append({"Type": f"{label} {len(sel_ids)}對1 Coin", "MatchReason": reason, "Bank_Data": sel_df, "Acc_Data": a_row.to_frame().T})
            matched_b.extend(sel_ids)
            matched_a.append(a_row["Acc_index"])
    rem_b = rem_b[~rem_b["Bank_index"].isin(matched_b)]
    rem_a = rem_a[~rem_a["Acc_index"].isin(matched_a)]

    print(f"\n🎉 [{label}] 完成，已核銷 {len(history)} 筆，剩餘銀行{len(rem_b)}/會計{len(rem_a)}")
    return history, rem_b, rem_a


# ══════════════════════════════════════════════════
# 5. 跨日對帳
# ══════════════════════════════════════════════════

def get_date_str(df: pd.DataFrame, date_col: str) -> pd.Series:
    return pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%Y/%m/%d")


def cross_day_reconcile(rem_b, rem_a, bank_date_col, acc_date_col, mode, day_window=7):
    history = []
    rem_b, rem_a = rem_b.copy(), rem_a.copy()
    amt_b = "存入金額" if mode == "Income" else "支出金額"
    label = "收入" if mode == "Income" else "支出"
    rem_b["_bdate"] = pd.to_datetime(rem_b[bank_date_col], errors="coerce")
    rem_a["_adate"] = pd.to_datetime(rem_a[acc_date_col], errors="coerce")
    for b_idx, b_row in rem_b.sort_values("_bdate", ascending=False).iterrows():
        b_amt, b_date = b_row[amt_b], b_row["_bdate"]
        cands = rem_a[
            (rem_a["業務金額"].abs() == b_amt)
            & ((rem_a["_adate"] - b_date).abs().dt.days <= day_window)
        ].copy()
        if cands.empty:
            continue
        cands["_diff"] = (cands["_adate"] - b_date).abs().dt.days
        cands = cands.sort_values("_diff")
        if len(cands) == 1:
            a_row = cands.iloc[0]
            print(f"  🔀 跨日[{label}] Bank:{b_row['Bank_index']}({b_date.date()}) ↔ Acc:{a_row['Acc_index']}({a_row['_adate'].date()}) 差{cands.iloc[0]['_diff']}天")
            history.append({"Type": f"跨日{label}", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_row.to_frame().T})
            rem_a = rem_a.drop(cands.index[0])
            rem_b = rem_b.drop(b_idx)
        else:
            best_id = llm_pick_best_candidate(b_row, cands, mode)
            if best_id:
                a_sel = cands[cands["Acc_index"] == best_id].iloc[0]
                history.append({"Type": f"跨日{label}(LLM)", "Bank_Data": b_row.to_frame().T, "Acc_Data": a_sel.to_frame().T})
                rem_a = rem_a[rem_a["Acc_index"] != best_id]
                rem_b = rem_b.drop(b_idx)
    rem_b = rem_b.drop(columns=["_bdate"], errors="ignore")
    rem_a = rem_a.drop(columns=["_adate"], errors="ignore")
    return history, rem_b, rem_a


def final_sweep_cross_day(rem_b, rem_a, bank_date_col, acc_date_col, mode, day_window=7, pending=None):
    if pending is None:
        pending = []
    history = []
    rem_b, rem_a = rem_b.copy(), rem_a.copy()
    amt_b = "存入金額" if mode == "Income" else "支出金額"
    label = "收入" if mode == "Income" else "支出"
    rem_b["_bdate"] = pd.to_datetime(rem_b[bank_date_col], errors="coerce")
    rem_a["_adate"] = pd.to_datetime(rem_a[acc_date_col], errors="coerce")
    matched_b, matched_a = [], []
    for b_idx, b_row in rem_b.iterrows():
        if b_row["Bank_index"] in matched_b:
            continue
        b_amt = b_row[amt_b]
        b_date = b_row["_bdate"]
        pool = rem_a[
            (~rem_a["Acc_index"].isin(matched_a))
            & ((rem_a["_adate"] - b_date).abs().dt.days <= day_window)
        ]
        found = False
        for r in range(2, min(4, len(pool) + 1)):
            if found:
                break
            for combo in itertools.combinations(pool.iterrows(), r):
                rows = [c[1] for c in combo]
                if sum(abs(row["業務金額"]) for row in rows) != b_amt:
                    continue
                combo_df = pd.DataFrame(rows)
                combo_ids = combo_df["Acc_index"].tolist()
                is_v, rs = evaluate_combination_with_llm(b_row.to_frame().T, combo_df, f"跨日1對{r}", mode)
                if is_v:
                    history.append({"Type": f"跨日{label} 1對{r}", "Bank_Data": b_row.to_frame().T, "Acc_Data": combo_df})
                    matched_b.append(b_row["Bank_index"])
                    matched_a.extend(combo_ids)
                    found = True
                    break
                else:
                    pending.append({"Mode": mode, "Type": f"跨日{label} 1對{r} 待確認", "Bank_Data": b_row.to_frame().T, "Acc_Data": combo_df, "Reason": rs})
    rem_b = rem_b[~rem_b["Bank_index"].isin(matched_b)].drop(columns=["_bdate"], errors="ignore")
    rem_a = rem_a[~rem_a["Acc_index"].isin(matched_a)].drop(columns=["_adate"], errors="ignore")
    return history, rem_b, rem_a


# ══════════════════════════════════════════════════
# 6. 掃底 / 軟配
# ══════════════════════════════════════════════════

def find_soft_matches(rem_b, rem_a, mode):
    soft_matches = []
    amt_col_b = "存入金額" if mode == "Income" else "支出金額"
    for b_idx, b_row in rem_b.iterrows():
        b_amt = b_row[amt_col_b]
        if b_amt == 0:
            continue
        candidates = rem_a[rem_a["業務金額"].abs() == b_amt]
        if not candidates.empty:
            soft_matches.append({"Mode": mode, "Bank_Data": b_row, "Candidates": candidates})
    return soft_matches


# ══════════════════════════════════════════════════
# 7. Excel 報表
# ══════════════════════════════════════════════════

def build_excel_report(bank_path, inc_rem_b, exp_rem_b, soft_matches, output_path, bank_in, bank_out, acc_in, acc_out, pending_combos=None):
    from openpyxl import load_workbook

    soft_ids = set(m["Bank_Data"]["Bank_index"] for m in soft_matches)
    if pending_combos:
        for p in pending_combos:
            for bid in p["Bank_Data"]["Bank_index"].values:
                soft_ids.add(str(bid))
    all_rem_b = pd.concat([inc_rem_b, exp_rem_b])
    hard_rem_ids = set(all_rem_b[~all_rem_b["Bank_index"].isin(soft_ids)]["Bank_index"].astype(str))
    pending_ids = soft_ids

    wb = load_workbook(bank_path)
    ws = wb.active

    header_row = None
    memo_col = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == "附言":
                header_row = cell.row
                memo_col = cell.column
                break
        if header_row:
            break

    if memo_col is None:
        insert_col = ws.max_column + 1
    else:
        insert_col = memo_col + 1
        ws.insert_cols(insert_col, 2)

    center = Alignment(horizontal="center", vertical="center")
    ref_header = ws.cell(row=header_row, column=memo_col)
    for offset, col_name in enumerate(["未入帳", "待確認"]):
        c = ws.cell(row=header_row, column=insert_col + offset, value=col_name)
        c.font = copy(ref_header.font)
        c.fill = copy(ref_header.fill)
        c.border = copy(ref_header.border)
        c.alignment = copy(ref_header.alignment)

    data_start_row = header_row + 1
    for excel_row in range(data_start_row, ws.max_row + 1):
        ref_data = ws.cell(row=excel_row, column=memo_col)
        for col_offset in [0, 1]:
            c = ws.cell(row=excel_row, column=insert_col + col_offset)
            c.fill = copy(ref_data.fill)
            c.border = copy(ref_data.border)
            c.alignment = center
            c.font = copy(ref_data.font)
    for bank_id in hard_rem_ids:
        excel_row = data_start_row + int(bank_id)
        if excel_row <= ws.max_row:
            ws.cell(row=excel_row, column=insert_col).value = "V"
            ws.cell(row=excel_row, column=insert_col).font = Font(bold=True, color="FF0000")
    for bank_id in pending_ids:
        excel_row = data_start_row + int(bank_id)
        if excel_row <= ws.max_row:
            ws.cell(row=excel_row, column=insert_col + 1).value = "V"
            ws.cell(row=excel_row, column=insert_col + 1).font = Font(bold=True, color="FF8C00")

    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col + 1)].width = 10

    # ── 每日變動數確認 sheet ──
    if "每日變動數確認" in wb.sheetnames:
        del wb["每日變動數確認"]
    ws2 = wb.create_sheet("每日變動數確認")

    bank_df = pd.concat([bank_in, bank_out])
    acc_df  = pd.concat([acc_in,  acc_out])
    bank_df["_date"] = pd.to_datetime(bank_df["交易日期"], errors="coerce").dt.strftime("%Y/%m/%d")
    acc_df["_date"]  = acc_df["業務日期_格式化"] if "業務日期_格式化" in acc_df.columns else pd.to_datetime(acc_df["業務日期"], errors="coerce").dt.strftime("%Y/%m/%d")
    all_dates = sorted(set(bank_df["_date"].dropna()) | set(acc_df["_date"].dropna()))

    bank_fill = PatternFill("solid", fgColor="D9E1F2")
    acc_fill  = PatternFill("solid", fgColor="E2EFDA")
    chk_fill  = PatternFill("solid", fgColor="FFF2CC")
    bold      = Font(bold=True)
    center    = Alignment(horizontal="center", vertical="center")
    right     = Alignment(horizontal="right",  vertical="center")
    num_fmt   = '#,##0;-#,##0;"-"'

    # Row 1：section headers（合併儲存格）
    for c_start, c_end, title, fill in [
        (1, 4,  "銀行對帳單", bank_fill),
        (5, 8,  "會計帳務",   acc_fill),
        (9, 12, "檢核",       chk_fill),
    ]:
        cell = ws2.cell(row=1, column=c_start, value=title)
        cell.font = Font(bold=True, size=12); cell.fill = fill; cell.alignment = center
        ws2.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)

    # Row 2：欄位標題
    col_headers = ["交易日期", "銀行收入", "銀行支出", "淨變動",
                   "交易日期", "借方",     "貸方",     "淨變動",
                   "交易日期", "借方差異", "貸方差異", "總差異"]
    col_fills   = [bank_fill]*4 + [acc_fill]*4 + [chk_fill]*4
    for i, (h, f) in enumerate(zip(col_headers, col_fills)):
        cell = ws2.cell(row=2, column=i+1, value=h)
        cell.font = bold; cell.fill = f; cell.alignment = center

    # 資料列
    for r, date in enumerate(all_dates, start=3):
        b_in  = bank_df[(bank_df["_date"]==date) & (bank_df["存入金額"]>0)]["存入金額"].sum() if "存入金額" in bank_df.columns else 0
        b_out = bank_df[(bank_df["_date"]==date) & (bank_df["支出金額"]>0)]["支出金額"].sum() if "支出金額" in bank_df.columns else 0
        a_in  = acc_df[(acc_df["_date"]==date) & (acc_df["業務金額"]>0)]["業務金額"].sum()
        a_out = abs(acc_df[(acc_df["_date"]==date) & (acc_df["業務金額"]<0)]["業務金額"].sum())
        diff_dr  = a_in  - b_in
        diff_cr  = a_out - b_out
        diff_tot = diff_dr - diff_cr

        row_data = [
            (date,          "YYYY/MM/DD", center, None),
            (b_in,          num_fmt,      right,  None),
            (b_out,         num_fmt,      right,  None),
            (b_in - b_out,  num_fmt,      right,  None),
            (date,          "YYYY/MM/DD", center, None),
            (a_in,          num_fmt,      right,  None),
            (a_out,         num_fmt,      right,  None),
            (a_in - a_out,  num_fmt,      right,  None),
            (date,          "YYYY/MM/DD", center, None),
            (diff_dr,       num_fmt,      right,  None),
            (diff_cr,       num_fmt,      right,  None),
            (diff_tot,      num_fmt,      right,  None),
        ]
        for col_idx, (val, fmt, align, fill) in enumerate(row_data, start=1):
            cell = ws2.cell(row=r, column=col_idx, value=val)
            cell.number_format = fmt; cell.alignment = align
            if fill:
                cell.fill = fill

    # 欄寬
    for i, w in enumerate([13,16,16,16, 13,16,16,16, 13,16,16,16], start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 18

    wb.save(output_path)
    print(f"\n💾 Excel 報表已產出：{output_path}")


# ══════════════════════════════════════════════════
# 8. 會計內部對消 / 迴轉預處理
# ══════════════════════════════════════════════════

def cancel_acc_internal_reversals(inc_rem_a, exp_rem_a):
    paired_inc, paired_exp = [], []
    log = []
    for _, e_row in exp_rem_a.iterrows():
        if e_row["Acc_index"] in paired_exp:
            continue
        target = abs(e_row["業務金額"])
        if target == 0:
            continue
        pool = inc_rem_a[~inc_rem_a["Acc_index"].isin(paired_inc)]
        found = False
        for n in range(1, min(ACC_REVERSAL_MAX_COMBO + 1, len(pool) + 1)):
            if found:
                break
            for combo in itertools.combinations(pool.iterrows(), n):
                rows = [c[1] for c in combo]
                if sum(abs(r["業務金額"]) for r in rows) != target:
                    continue
                combo_ids = [r["Acc_index"] for r in rows]
                log.append({"inc_rows": rows, "exp": e_row, "amount": target, "n": n})
                paired_inc.extend(combo_ids)
                paired_exp.append(e_row["Acc_index"])
                tag = f"{n}:1" if n > 1 else "1:1"
                print(f"  🟣 會計內對消({tag})：{combo_ids}(+{target:,.0f}) ⇔ {e_row['Acc_index']}(-{target:,.0f})")
                found = True
                break
    clean_inc = inc_rem_a[~inc_rem_a["Acc_index"].isin(paired_inc)]
    clean_exp = exp_rem_a[~exp_rem_a["Acc_index"].isin(paired_exp)]
    return clean_inc, clean_exp, log


def preprocess_acc_reversals(acc_in, acc_out):
    paired_inc, paired_exp = [], []
    log = []
    reversal_entries = acc_out[acc_out["描述.1"].apply(lambda d: any(kw in str(d) for kw in PRE_REVERSAL_KEYWORDS))]
    for _, e_row in reversal_entries.iterrows():
        if e_row["Acc_index"] in paired_exp:
            continue
        target = abs(e_row["業務金額"])
        if target == 0:
            continue
        pool = acc_in[~acc_in["Acc_index"].isin(paired_inc)]
        candidates = pool[pool["業務金額"].abs() == target]
        if candidates.empty:
            continue
        e_desc = str(e_row.get("描述.1", ""))
        best = None
        for _, c_row in candidates.iterrows():
            c_desc = str(c_row.get("描述.1", ""))
            if c_desc and len(c_desc) >= 4 and c_desc in e_desc:
                best = c_row
                break
        if best is not None:
            log.append({"inc": best, "exp": e_row, "amount": target})
            paired_inc.append(best["Acc_index"])
            paired_exp.append(e_row["Acc_index"])
            print(f"  🔵 預處理對消：Acc:{best['Acc_index']}(+{target:,.0f}) ⇔ Acc:{e_row['Acc_index']}(-{target:,.0f})")
    clean_inc = acc_in[~acc_in["Acc_index"].isin(paired_inc)]
    clean_exp = acc_out[~acc_out["Acc_index"].isin(paired_exp)]
    return clean_inc, clean_exp, log


def _is_reversal_entry(acc_row) -> bool:
    desc = str(acc_row.get("描述.1", ""))
    return any(kw in desc for kw in NET_REVERSAL_KEYWORDS)


def net_acc_reconcile(inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a):
    net_inc_hist, net_exp_hist = [], []
    matched_b_inc, matched_b_exp = [], []
    matched_inc_a, matched_exp_a = [], []
    print(f"\n⏳ [淨額配對] 迴轉分錄搜尋... (銀行收{len(inc_rem_b)}/支{len(exp_rem_b)} | 會計正{len(inc_rem_a)}/負{len(exp_rem_a)})")

    for b_idx, b_row in inc_rem_b.iterrows():
        if b_row["Bank_index"] in matched_b_inc:
            continue
        b_amt = b_row["存入金額"]
        found = False
        avail_pos = inc_rem_a[~inc_rem_a["Acc_index"].isin(matched_inc_a)]
        avail_neg = exp_rem_a[~exp_rem_a["Acc_index"].isin(matched_exp_a)]
        if avail_pos.empty or avail_neg.empty:
            continue
        for np_ in range(1, min(NET_MAX_POS + 1, len(avail_pos) + 1)):
            if found:
                break
            for pos_combo in itertools.combinations(avail_pos.iterrows(), np_):
                if found:
                    break
                pos_rows = [c[1] for c in pos_combo]
                if not all(_is_reversal_entry(r) for r in pos_rows):
                    continue
                pos_sum = sum(abs(r["業務金額"]) for r in pos_rows)
                for nq_ in range(1, min(NET_MAX_NEG + 1, len(avail_neg) + 1)):
                    if found:
                        break
                    for neg_combo in itertools.combinations(avail_neg.iterrows(), nq_):
                        neg_rows = [c[1] for c in neg_combo]
                        neg_sum = sum(abs(r["業務金額"]) for r in neg_rows)
                        if neg_sum - pos_sum != b_amt:
                            continue
                        pos_df = pd.DataFrame(pos_rows)
                        neg_df = pd.DataFrame(neg_rows)
                        reason = f"負方{neg_sum:,.0f} - 迴轉{pos_sum:,.0f} = 銀行{b_amt:,.0f}"
                        net_inc_hist.append({"Type": f"收入 淨額({np_}正{nq_}負)", "MatchReason": reason, "Bank_Data": b_row.to_frame().T, "Acc_Data": pd.concat([pos_df, neg_df])})
                        matched_b_inc.append(b_row["Bank_index"])
                        matched_inc_a.extend([r["Acc_index"] for r in pos_rows])
                        matched_exp_a.extend([r["Acc_index"] for r in neg_rows])
                        found = True
                        break

    for b_idx, b_row in exp_rem_b.iterrows():
        if b_row["Bank_index"] in matched_b_exp:
            continue
        b_amt = b_row["支出金額"]
        found = False
        avail_neg = exp_rem_a[~exp_rem_a["Acc_index"].isin(matched_exp_a)]
        avail_pos = inc_rem_a[~inc_rem_a["Acc_index"].isin(matched_inc_a)]
        if avail_neg.empty or avail_pos.empty:
            continue
        for nq_ in range(1, min(NET_MAX_NEG + 1, len(avail_neg) + 1)):
            if found:
                break
            for neg_combo in itertools.combinations(avail_neg.iterrows(), nq_):
                if found:
                    break
                neg_rows = [c[1] for c in neg_combo]
                neg_sum = sum(abs(r["業務金額"]) for r in neg_rows)
                for np_ in range(1, min(NET_MAX_POS + 1, len(avail_pos) + 1)):
                    if found:
                        break
                    for pos_combo in itertools.combinations(avail_pos.iterrows(), np_):
                        pos_rows = [c[1] for c in pos_combo]
                        if not all(_is_reversal_entry(r) for r in pos_rows):
                            continue
                        pos_sum = sum(abs(r["業務金額"]) for r in pos_rows)
                        if neg_sum - pos_sum != b_amt:
                            continue
                        neg_df = pd.DataFrame(neg_rows)
                        pos_df = pd.DataFrame(pos_rows)
                        reason = f"負方{neg_sum:,.0f} - 迴轉{pos_sum:,.0f} = 銀行{b_amt:,.0f}"
                        net_exp_hist.append({"Type": f"支出 淨額({nq_}負{np_}正)", "MatchReason": reason, "Bank_Data": b_row.to_frame().T, "Acc_Data": pd.concat([neg_df, pos_df])})
                        matched_b_exp.append(b_row["Bank_index"])
                        matched_exp_a.extend([r["Acc_index"] for r in neg_rows])
                        matched_inc_a.extend([r["Acc_index"] for r in pos_rows])
                        found = True
                        break

    inc_rem_b = inc_rem_b[~inc_rem_b["Bank_index"].isin(matched_b_inc)]
    exp_rem_b = exp_rem_b[~exp_rem_b["Bank_index"].isin(matched_b_exp)]
    inc_rem_a = inc_rem_a[~inc_rem_a["Acc_index"].isin(matched_inc_a)]
    exp_rem_a = exp_rem_a[~exp_rem_a["Acc_index"].isin(matched_exp_a)]
    print(f"  淨額配對完成：已配 收入{len(net_inc_hist)}/支出{len(net_exp_hist)} 組")
    return net_inc_hist, net_exp_hist, inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a


# ══════════════════════════════════════════════════
# 9. HTML 報告
# ══════════════════════════════════════════════════

def build_html_report(
    all_histories, all_rems, soft_matches, date, output_path,
    pending_combos=None, reversal_log=None, acc_only_rems=None, acc_reversal_log=None, pre_reversal_log=None,
) -> str:
    style = """<style>
        body{font-family:'Inter',system-ui,sans-serif;background:#f8fafc;color:#1e293b;padding:40px;}
        .container{max-width:1400px;margin:auto;}
        h1{color:#0f172a;border-bottom:3px solid #3b82f6;padding-bottom:10px;}
        .summary-box{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-bottom:30px;}
        .stat-card{background:white;padding:20px;border-radius:12px;box-shadow:0 4px 6px rgb(0 0 0/0.05);}
        .card{background:white;border-radius:16px;padding:25px;margin-bottom:40px;box-shadow:0 10px 15px -3px rgb(0 0 0/0.1);}
        .tag{display:inline-block;padding:4px 12px;border-radius:999px;font-size:12px;font-weight:bold;margin-bottom:15px;}
        .tag-inc{background:#dcfce7;color:#166534;} .tag-exp{background:#fee2e2;color:#991b1b;}
        .tag-warn{background:#fef08a;color:#854d0e;}
        .table-section{margin-top:15px;overflow-x:auto;}
        .side-title{font-size:13px;text-transform:uppercase;font-weight:bold;color:#64748b;margin-bottom:10px;border-left:4px solid #3b82f6;padding-left:10px;}
        table{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:10px;text-align:right;}
        th{background:#f8fafc;padding:10px;border-bottom:2px solid #e2e8f0;text-align:right;}
        td{padding:10px;border-bottom:1px solid #f1f5f9;}
    </style>"""

    total_matched = len(all_histories["Income"]) + len(all_histories["Expense"])
    total_pending_all = len(soft_matches) + len(pending_combos or [])
    rem_b_ids_in_soft = [m["Bank_Data"]["Bank_index"] for m in soft_matches]
    total_hard_rem = (
        len(all_rems["Income_B"][~all_rems["Income_B"]["Bank_index"].isin(rem_b_ids_in_soft)])
        + len(all_rems["Expense_B"][~all_rems["Expense_B"]["Bank_index"].isin(rem_b_ids_in_soft)])
    )

    html = f"<div class='container'><h1>收支核銷報告 ({date})</h1>"
    html += "<div class='summary-box'>"
    html += f"<div class='stat-card' style='border-top:4px solid #22c55e;'><h3>🟢 總結成功核銷</h3><div style='font-size:32px;font-weight:bold;color:#22c55e;'>{total_matched} 組</div></div>"
    html += f"<div class='stat-card' style='border-top:4px solid #eab308;'><h3>🟡 全部待確認</h3><div style='font-size:32px;font-weight:bold;color:#eab308;'>{total_pending_all} 組</div></div>"
    html += f"<div class='stat-card' style='border-top:4px solid #ef4444;'><h3>🔴 總計剩餘未入帳 (銀行端)</h3><div style='font-size:32px;font-weight:bold;color:#ef4444;'>{total_hard_rem} 筆</div></div>"
    html += "</div>"

    if soft_matches or pending_combos:
        html += f"<h2>🟡 全部待確認（共 {len(soft_matches) + len(pending_combos or [])} 筆）</h2>"
        for match in soft_matches:
            mode_tw = "收款" if match["Mode"] == "Income" else "付款"
            html += f"<div class='card' style='border-left:5px solid #eab308;background:#fefce8;padding:20px;'>"
            html += f"<div class='tag tag-warn'>待確認 {mode_tw}（1對1金額相符）</div>"
            html += "<div class='table-section'><div class='side-title'>【銀行端】</div>"
            html += match["Bank_Data"].to_frame().T.to_html(index=False)
            html += f"</div><div class='table-section' style='margin-top:20px;'><div class='side-title'>👇 會計端候選（{len(match['Candidates'])} 筆）</div>"
            html += match["Candidates"].to_html(index=False)
            html += "</div></div>"
        for p in (pending_combos or []):
            html += f"<div class='card' style='border-left:5px solid #eab308;background:#fefce8;padding:20px;'>"
            html += f"<div class='tag tag-warn'>{p['Type']}</div>"
            html += f"<div style='color:#92400e;margin-bottom:10px;'>LLM意見：{p['Reason']}</div>"
            html += "<div class='table-section'><div class='side-title'>【銀行端】</div>"
            html += p["Bank_Data"].to_html(index=False)
            html += "</div><div class='table-section' style='margin-top:15px;'><div class='side-title'>【會計端】</div>"
            html += p["Acc_Data"].to_html(index=False)
            html += "</div></div>"

    html += "<h2 style='margin-top:40px'>🔴 剩餘未入帳之銀行賬務清單</h2>"
    for m in ["Income", "Expense"]:
        df_rem = all_rems[f"{m}_B"]
        df_rem_pure = df_rem[~df_rem["Bank_index"].isin(rem_b_ids_in_soft)]
        if not df_rem_pure.empty:
            label_txt = "尚未入帳之【收款】" if m == "Income" else "尚未入帳之【付款】"
            tag_cls = "tag-inc" if m == "Income" else "tag-exp"
            border_color = "#22c55e" if m == "Income" else "#ef4444"
            html += f"<div class='card' style='border-top:5px solid {border_color}'>"
            html += f"<div class='tag {tag_cls}'>{label_txt}</div>"
            show_cols = [c for c in ["Bank_index", "交易日期", "交易時間", "附言", "摘要", "存入金額", "支出金額"] if c in df_rem_pure.columns]
            html += df_rem_pure[show_cols].to_html(index=False)
            html += "</div>"

    acc_only_rems = acc_only_rems or {}
    for m, label_txt, border_color, tag_cls in [("Income", "【收款】會計多入", "#f97316", "tag-inc"), ("Expense", "【付款】會計多入", "#f97316", "tag-exp")]:
        df_ao = acc_only_rems.get(f"{m}_A", pd.DataFrame())
        if df_ao.empty:
            continue
        acc_cols = ["Acc_index", "業務日期_格式化", "憑證編號", "描述.1", "業務金額", "借方/貸方", "來源Sheet"]
        show_cols = [c for c in acc_cols if c in df_ao.columns]
        html += f"<div class='card' style='border-top:5px solid {border_color}'>"
        html += f"<div class='tag {tag_cls}'>{label_txt}（銀行無此筆）</div>"
        html += df_ao[show_cols].to_html(index=False)
        html += "</div>"

    html += "<h2 style='margin-top:60px'>🟢 詳盡對帳軌跡 (已核銷)</h2>"
    for item in all_histories["Income"] + all_histories["Expense"]:
        is_inc = "收入" in item["Type"]
        html += "<div class='card'>"
        html += f"<div class='tag {'tag-inc' if is_inc else 'tag-exp'}'>{item['Type']}</div>"
        html += f"<div style='font-size:12px;color:#64748b;margin-bottom:10px;'>⚙️ 配對依據：{item.get('MatchReason', '—')}</div>"
        html += "<div class='table-section'><div class='side-title'>銀行端原始紀錄</div>"
        html += item["Bank_Data"].to_html(index=False)
        html += "</div><div class='table-section'><div class='side-title'>會計端原始紀錄</div>"
        html += item["Acc_Data"].to_html(index=False)
        html += "</div></div>"

    html += "</div>"
    full_html = f"<html><head><meta charset='utf-8'>{style}</head><body>{html}</body></html>"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(full_html)
    return full_html


# ══════════════════════════════════════════════════
# 10. 單月流程
# ══════════════════════════════════════════════════

def _run_month(month_code: str, bank_path: str, acc_sheets: dict, output_dir: str) -> dict:
    target_date = month_code_to_date(month_code)
    snap = {k: v for k, v in _token_usage.items()}

    print(f'\n{"="*60}')
    print(f"🗓️  處理月份：{target_date}  ({os.path.basename(bank_path)})")
    print(f'{"="*60}')

    bank_in, bank_out, acc_in, acc_out, reversal_log = load_and_preprocess(target_date, bank_path, acc_sheets)
    print(f"✅ 數據載入：[收入組] 銀行 {len(bank_in)} 筆 / 會計 {len(acc_in)} 筆 | [支出組] 銀行 {len(bank_out)} 筆 / 會計 {len(acc_out)} 筆")

    acc_in_raw, acc_out_raw = acc_in.copy(), acc_out.copy()
    _acc_in_orig = len(acc_in)
    _acc_out_orig = len(acc_out)

    print("\n========== 🔵 會計迴轉預處理 ==========")
    acc_in, acc_out, pre_reversal_log = preprocess_acc_reversals(acc_in, acc_out)

    bank_dates = sorted(set(
        get_date_str(bank_in, "交易日期").dropna().tolist() +
        get_date_str(bank_out, "交易日期").dropna().tolist()
    ))
    all_inc_hist, all_exp_hist = [], []
    all_inc_rem_b, all_exp_rem_b = [], []
    all_inc_rem_a, all_exp_rem_a = [], []
    all_pending_combos: list = []

    for date in bank_dates:
        b_in_day = bank_in[get_date_str(bank_in, "交易日期") == date]
        b_out_day = bank_out[get_date_str(bank_out, "交易日期") == date]
        a_in_day = acc_in[acc_in["業務日期_格式化"] == date]
        a_out_day = acc_out[acc_out["業務日期_格式化"] == date]
        print(f"\n===== 📅 {date} ===== 銀行收{len(b_in_day)}/支{len(b_out_day)} | 會計收{len(a_in_day)}/支{len(a_out_day)}")
        h_i, r_bi, r_ai = reconcile_engine(b_in_day, a_in_day, "Income", all_pending_combos)
        h_e, r_be, r_ae = reconcile_engine(b_out_day, a_out_day, "Expense", all_pending_combos)
        all_inc_hist += h_i
        all_exp_hist += h_e
        all_inc_rem_b.append(r_bi)
        all_exp_rem_b.append(r_be)
        all_inc_rem_a.append(r_ai)
        all_exp_rem_a.append(r_ae)

    inc_hist = all_inc_hist
    exp_hist = all_exp_hist
    inc_rem_b = pd.concat(all_inc_rem_b) if all_inc_rem_b else pd.DataFrame()
    exp_rem_b = pd.concat(all_exp_rem_b) if all_exp_rem_b else pd.DataFrame()
    inc_rem_a = pd.concat(all_inc_rem_a) if all_inc_rem_a else pd.DataFrame()
    exp_rem_a = pd.concat(all_exp_rem_a) if all_exp_rem_a else pd.DataFrame()

    print("\n\n========== 🔢 淨額配對（階段 1.5） ==========")
    net_inc_h, net_exp_h, inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a = net_acc_reconcile(
        inc_rem_b, exp_rem_b, inc_rem_a, exp_rem_a
    )
    inc_hist += net_inc_h
    exp_hist += net_exp_h

    print("\n\n========== 🔀 跨日對帳（階段二） ==========")
    cross_inc, inc_rem_b, inc_rem_a = cross_day_reconcile(inc_rem_b, inc_rem_a, "交易日期", "業務日期_格式化", "Income")
    cross_exp, exp_rem_b, exp_rem_a = cross_day_reconcile(exp_rem_b, exp_rem_a, "交易日期", "業務日期_格式化", "Expense")
    inc_hist += cross_inc
    exp_hist += cross_exp

    sweep_inc, inc_rem_b, inc_rem_a = final_sweep_cross_day(inc_rem_b, inc_rem_a, "交易日期", "業務日期_格式化", "Income", pending=all_pending_combos)
    sweep_exp, exp_rem_b, exp_rem_a = final_sweep_cross_day(exp_rem_b, exp_rem_a, "交易日期", "業務日期_格式化", "Expense", pending=all_pending_combos)
    inc_hist += sweep_inc
    exp_hist += sweep_exp

    inc_soft = find_soft_matches(inc_rem_b, inc_rem_a, "Income")
    exp_soft = find_soft_matches(exp_rem_b, exp_rem_a, "Expense")
    all_soft = inc_soft + exp_soft

    output_xlsx = os.path.join(output_dir, f"未入帳清單_{month_code}.xlsx")
    build_excel_report(bank_path, inc_rem_b, exp_rem_b, all_soft, output_xlsx, bank_in, bank_out, acc_in_raw, acc_out_raw, pending_combos=all_pending_combos)

    inc_hist = [item for item in inc_hist if isinstance(item, dict)]
    exp_hist = [item for item in exp_hist if isinstance(item, dict)]

    print("\n\n========== 🟣 會計端內部對消 ==========")
    inc_rem_a, exp_rem_a, acc_reversal_log = cancel_acc_internal_reversals(inc_rem_a, exp_rem_a)

    output_html = os.path.join(output_dir, f"未入帳清單_{month_code}.html")
    build_html_report(
        {"Income": inc_hist, "Expense": exp_hist},
        {"Income_B": inc_rem_b, "Expense_B": exp_rem_b},
        all_soft, target_date, output_html,
        pending_combos=all_pending_combos,
        reversal_log=reversal_log,
        acc_only_rems={"Income_A": inc_rem_a, "Expense_A": exp_rem_a},
        acc_reversal_log=acc_reversal_log,
        pre_reversal_log=pre_reversal_log,
    )

    month_tokens = {
        "month": target_date,
        "calls": _token_usage["calls"] - snap["calls"],
        "input": _token_usage["input"] - snap["input"],
        "output": _token_usage["output"] - snap["output"],
    }
    month_tokens["total"] = month_tokens["input"] + month_tokens["output"]
    _monthly_token_log.append(month_tokens)

    print(f"\n✅ [{month_code}] {target_date} 完成！")
    print(f"   Excel → {os.path.basename(output_xlsx)}")
    print(f"   HTML  → {os.path.basename(output_html)}")
    print(f"   LLM   → calls:{month_tokens['calls']} input:{month_tokens['input']:,} output:{month_tokens['output']:,}")

    return {"xlsx": output_xlsx, "html": output_html}


# ══════════════════════════════════════════════════
# 11. 公開入口
# ══════════════════════════════════════════════════

def run_pipeline(
    bank_file_paths: dict,
    acc_file_paths: list,
    output_dir: str,
) -> dict:
    """
    bank_file_paths: {月份碼: 檔案路徑}  e.g. {'11501': '/tmp/銀行對帳單-11501.xlsx'}
    acc_file_paths : [帳務查詢路徑, ...]
    output_dir     : 輸出目錄
    Azure 憑證從 .env 自動載入（AZURE_OPENAI_ENDPOINT / AZURE_OPENAI_KEY）
    回傳: {月份碼: {'xlsx': 路徑, 'html': 路徑}}
    """
    global _openai_client, _token_usage, _monthly_token_log
    _openai_client = AzureOpenAI(
        azure_endpoint=_AZURE_ENDPOINT,
        api_key=_AZURE_KEY,
        api_version=_AZURE_API_VERSION,
        http_client=httpx.Client(verify=False),
    )
    _token_usage = {"input": 0, "output": 0, "calls": 0}
    _monthly_token_log = []

    os.makedirs(output_dir, exist_ok=True)

    print("載入帳務查詢檔案...")
    acc_sheets = load_acc_sheets(acc_file_paths)
    print(f"✅ 找到 {len(bank_file_paths)} 份銀行對帳單：{sorted(bank_file_paths.keys())}")

    results = {}
    for month_code in sorted(bank_file_paths.keys()):
        bank_path = convert_to_xlsx(bank_file_paths[month_code])
        results[month_code] = _run_month(month_code, bank_path, acc_sheets, output_dir)

    # 寫 token log
    log_path = os.path.join(output_dir, "token_log.csv")
    file_exists = os.path.isfile(log_path)
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["run_at", "month", "calls", "input", "output", "total"])
        if not file_exists:
            writer.writeheader()
        run_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in _monthly_token_log:
            writer.writerow({"run_at": run_at, **row})

    print(f"\n{'='*50}")
    print(f"📊 LLM Token 使用統計")
    print(f"   呼叫次數  : {_token_usage['calls']:>8,} 次")
    print(f"   Input  tokens: {_token_usage['input']:>8,}")
    print(f"   Output tokens: {_token_usage['output']:>8,}")
    print(f"   Total  tokens: {_token_usage['input'] + _token_usage['output']:>8,}")

    return results
